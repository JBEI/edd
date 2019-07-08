# coding: utf-8
"""
Integration tests for ICE.
"""

from io import BytesIO

from django.test import tag
from django.urls import reverse
from openpyxl.workbook import Workbook
from requests import codes

from edd import TestCase
from jbei.rest.auth import HmacAuth
from jbei.rest.clients.ice import IceApi

from .. import models
from . import factory


def user_to_ice_json(user):
    # these fields are all required for ICE
    return {
        "firstName": user.first_name,
        "lastName": user.last_name,
        "initials": user.initials,
        "description": "",
        "institution": "",
        "email": user.email,
        "password": "",
    }


def ice_url(path):
    if path[0] == "/":
        path = path[1:]
    return f"http://ice:8080/rest/{path}"


class IceUserCheck(object):
    user_url = ice_url("/users")

    def __init__(self, ice):
        self.ice = ice

    def create(self, user):
        """Create a user and return its ID."""
        try:
            response = self.ice.session.post(
                self.user_url,
                json=user_to_ice_json(user),
                params={"sendEmail": "false"},
            )
            created = response.json()
            return created["id"]
        except ValueError as e:
            raise AssertionError(f"Bad response: {response.content}") from e
        except Exception as e:
            raise AssertionError(f"Failed to create user") from e

    def exists(self, user):
        """Return a user ID if exists, otherwise return None."""
        try:
            response = self.ice.session.get(
                self.user_url, params={"filter": user.email}
            )
            info = response.json()
            if info["resultCount"] != 0:
                return info["users"][0]["id"]
            return None
        except ValueError as e:
            raise AssertionError(f"Bad response: {response.content}") from e
        except Exception as e:
            raise AssertionError(f"Failed to check user") from e


@tag("integration")
class IceIntegrationTests(TestCase):
    """Sets of tests to validate communication between EDD and ICE."""

    @classmethod
    def setUpClass(cls):
        super(IceIntegrationTests, cls).setUpClass()
        auth = HmacAuth("edd", "Administrator")
        ice = IceApi(auth)
        try:
            # make sure ICE has users matching EDD users
            cls._ensureUsers(ice)
            # set the admin account type on admin_ice_user
            acct = user_to_ice_json(cls.admin_ice_user)
            acct.update(accountType="ADMIN")
            ice.session.put(ice_url(f"/users/{cls.admin_ice_user._ice_id}"), json=acct)
            # populate ICE with some strains
            with factory.load_test_file("ice_entries.csv") as entries:
                response = ice.session.post(
                    ice_url("/uploads/file"), files={"type": "strain", "file": entries}
                )
            upload_id = response.json()["uploadInfo"]["id"]
            response = ice.session.put(
                ice_url(f"/uploads/{upload_id}/status"),
                json={"id": upload_id, "status": "APPROVED"},
            )
            # fetch the part IDs
            response = ice.session.get(
                ice_url("/collections/available/entries"),
                params={"sort": "created", "asc": "false"},
            )
            entries = response.json()["data"][:10]
            cls.part_ids = [p["partId"] for p in reversed(entries)]
            cls.db_ids = [p["id"] for p in reversed(entries)]
            # set read permissions on some of the created strains
            for idx in range(5):
                response = ice.session.post(
                    ice_url(f"/parts/{cls.part_ids[idx]}/permissions"),
                    json={
                        "article": "ACCOUNT",
                        "articleId": cls.read_ice_user._ice_id,
                        "type": "READ_ENTRY",
                        "typeId": cls.db_ids[idx],
                    },
                )
        except Exception as e:
            cls.tearDownClass()
            raise e

    @classmethod
    def setUpTestData(cls):
        super(IceIntegrationTests, cls).setUpTestData()
        cls.admin_ice_user = factory.UserFactory(email="admin@example.org")
        cls.read_ice_user = factory.UserFactory(email="reader@example.org")
        cls.none_ice_user = factory.UserFactory(email="none@example.org")

    @classmethod
    def _ensureUsers(cls, ice):
        check = IceUserCheck(ice)
        for user in [cls.admin_ice_user, cls.read_ice_user, cls.none_ice_user]:
            user_id = check.exists(user)
            if user_id is None:
                user_id = check.create(user)
            user._ice_id = user_id

    def test_read_parts(self):
        admin_auth = HmacAuth("edd", self.admin_ice_user.email)
        reader_auth = HmacAuth("edd", self.read_ice_user.email)
        none_auth = HmacAuth("edd", self.none_ice_user.email)
        # verify that admin user finds all parts
        entries_url = ice_url("/collections/available/entries?sort=created&asc=false")
        ice = IceApi(admin_auth)
        response = ice.session.get(entries_url)
        payload = response.json()["data"]
        entries = {p["partId"] for p in payload}
        self.assertTrue(entries.issuperset(self.part_ids))
        # verify that reader user finds the five parts with permissions set
        ice = IceApi(reader_auth)
        response = ice.session.get(entries_url)
        payload = response.json()["data"]
        entries = {p["partId"] for p in payload}
        self.assertTrue(entries.issuperset(self.part_ids[:5]))
        self.assertEqual(len(entries.intersection(self.part_ids[5:])), 0)
        # verify that user with no permissions finds no parts
        ice = IceApi(none_auth)
        response = ice.session.get(entries_url)
        payload = response.json()["data"]
        entries = {p["partId"] for p in payload}
        self.assertEqual(len(entries.intersection(self.part_ids)), 0)

    def test_upload_links_admin(self):
        admin_study = factory.StudyFactory()
        admin_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.admin_ice_user
        )
        response = self._run_upload(self.part_ids, admin_study, self.admin_ice_user)
        # should return OK from upload
        self.assertEqual(response.status_code, codes.ok)
        # there should be 10 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=admin_study).distinct().count(), 10
        )
        # TODO: cannot check links because tests in transaction that ultimately calls rollback()
        # Celery task is only ever submitted when the transaction calls commit() successfully

    def test_upload_links_reader(self):
        reader_study = factory.StudyFactory()
        reader_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.read_ice_user
        )
        # should return 500 error on uploading admin-only strains
        # skip testing this until ICE-90 is resolved
        # response = self._run_upload(self.part_ids, reader_study, self.read_ice_user)
        # self.assertEqual(response.status_code, codes.server_error)
        # should return OK on uploading readable strains
        response = self._run_upload(self.part_ids[:5], reader_study, self.read_ice_user)
        self.assertEqual(response.status_code, codes.ok)
        # there should be 5 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=reader_study).distinct().count(), 5
        )
        # TODO: cannot check links because tests in transaction that ultimately calls rollback()
        # Celery task is only ever submitted when the transaction calls commit() successfully

    def test_upload_links_none(self):
        none_study = factory.StudyFactory()
        none_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.none_ice_user
        )
        # should return 500 error on uploading admin-only strains
        response = self._run_upload(self.part_ids, none_study, self.none_ice_user)
        self.assertEqual(response.status_code, codes.server_error)
        # should return 500 error on uploading reader-only strains
        response = self._run_upload(self.part_ids[:5], none_study, self.none_ice_user)
        self.assertEqual(response.status_code, codes.server_error)
        # there should be 0 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=none_study).distinct().count(), 0
        )

    def _create_workbook(self, parts):
        upload = BytesIO()
        wb = Workbook()
        ws = wb.active
        for i, title in enumerate(["Line Name", "Part ID", "Media"], 1):
            ws.cell(row=1, column=i).value = title
        for i, part in enumerate(parts, 2):
            ws.cell(row=i, column=1).value = part
            ws.cell(row=i, column=2).value = part
            ws.cell(row=i, column=3).value = "M9"
        wb.save(upload)
        upload.name = "description.xlsx"
        upload.content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        upload.seek(0)
        return upload

    def _run_upload(self, parts, study, user):
        upload = self._create_workbook(parts)
        self.client.force_login(user)
        response = self.client.post(
            reverse("main:describe", kwargs={"slug": study.slug}), data={"file": upload}
        )
        return response
