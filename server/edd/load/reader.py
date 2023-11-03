import csv
import logging
import typing
import uuid
from collections.abc import Iterable

from openpyxl import load_workbook

from . import exceptions

if typing.TYPE_CHECKING:
    Any = typing.Any
    from .layout import Cell, ImportLayout, ImportReader, Record, Row, Sheet

logger = logging.getLogger(__name__)
EMPTY_ROW: tuple[int, "Row"] = (-1, [])


class ExcelImportReader:
    """Reader for a Parser to use to interpret Excel payloads."""

    def __init__(self, *, multisheet=False):
        self.multisheet = multisheet

    def read(self, stream: "Any", layout: "ImportLayout") -> "Sheet":
        workbook = load_workbook(stream, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            layout.sheet(worksheet.title)
            yield from enumerate(worksheet.iter_rows())
            # giving an empty row to indicate we're done with the sheet
            yield EMPTY_ROW
            if not self.multisheet:
                raise exceptions.IgnoredWorksheetWarning(
                    processed_title=worksheet.title,
                    ignored_sheet_count=len(workbook.worksheets) - 1,
                )

    def value(self, cell: "Cell") -> "Any":
        value = cell.value
        if isinstance(value, str):
            return value.strip().strip("\ufeff")
        return value


class CsvImportReader:
    """Reader for a Parser to use to interpret CSV payloads."""

    def read(self, stream: "Any", layout: "ImportLayout") -> "Sheet":
        # generator from the reader
        yield from enumerate(csv.reader(stream))
        # add an empty row to indicate we're done
        yield EMPTY_ROW

    def value(self, cell: "Cell") -> str:
        return cell.strip().strip("\ufeff")


class Parser:
    def __init__(
        self,
        reader: "ImportReader",
        layout_class: type["ImportLayout"],
        *,
        load_uuid: uuid.UUID | None = None,
    ):
        self.reader = reader
        self.layout_class = layout_class
        self.uuid = load_uuid or uuid.uuid4()

    def consume_stream(
        self,
        stream: "Any",
        layout: "ImportLayout",
    ) -> Iterable["Record"]:
        for i, row in self.reader.read(stream, layout):
            try:
                yield from layout.process_row(row, i)
            except exceptions.EDDImportWarning as w:
                logger.warning("Problem parsing import row", exc_info=w)

    def layout(self) -> "ImportLayout":
        return self.layout_class(self)

    def parse(self, stream: "Any") -> Iterable["Record"]:
        layout = self.layout()
        try:
            yield from self.consume_stream(stream, layout)
        except exceptions.EDDImportError as e:
            logger.error("Error parsing import stream", exc_info=e)
            raise e
        except exceptions.EDDImportWarning as w:
            logger.warning("Problem parsing import stream", exc_info=w)
        finally:
            layout.finish()
