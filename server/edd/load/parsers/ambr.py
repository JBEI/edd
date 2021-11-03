import logging
from uuid import UUID

import numpy as np
from openpyxl import Workbook

from edd.load.models import DefaultUnit, MeasurementNameTransform
from main.models import MeasurementType

from .core import MultiSheetExcelParserMixin
from .generic import GenericImportParser

logger = logging.getLogger(__name__)


class AmbrExcelParser(MultiSheetExcelParserMixin, GenericImportParser):
    def __init__(self, import_uuid: UUID):
        super().__init__(import_uuid=import_uuid,)
        # self.parsed_result = pd.DataFrame()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def _parse_sheet_rows(self, name, sheet):

        # for every two columns in the worksheet
        # corresponding to each measurement type in the sheet
        for index in range(1, sheet.max_column + 1, 2):
            for col in sheet.iter_cols(min_row=1, min_col=index, max_col=index):
                time_data = [el.value for el in col]
            for col in sheet.iter_cols(min_row=1, min_col=index + 1, max_col=index + 1):
                mes_data = [el.value for el in col]

            # decimate the data here
            # check if data has more than 200 points then decimate else do not
            if len(mes_data) > 200:
                time_data = time_data[0::10]
                mes_data = mes_data[0::10]

            # using mapper to map data into the EDD import format
            # and convert in to a pandas dataframe
            # set the line name and the dataframe with the two columns
            # with data for the next measurement type
            self.map_data(name, (time_data, mes_data))

    def map_data(self, name, data):

        time_data, mes_data = data
        mtype_name = mes_data[0]

        try:
            # get EDD name for current measurement if mapping exists
            mes_transform_qs = MeasurementNameTransform.objects.all().filter(
                input_type_name=mtype_name, parser="ambr"
            )
            mtype_name = mes_transform_qs[0].edd_type_name.type_name
        except Exception as ex:
            logger.debug(
                f"Error trying to retrieve measurement type mapping \
                between ambr type and expected edd type name for {mtype_name}"
            )
            logger.exception(ex)

        try:
            # get default unit record for current measurement type
            mes_type_qs = MeasurementType.objects.all().filter(type_name=mtype_name)
        except Exception as ex:
            logger.debug(
                f"Error trying to retrieve measurement type \
            for {mtype_name}"
            )
            logger.exception(ex)

        try:
            du_qs = DefaultUnit.objects.all().filter(
                measurement_type=mes_type_qs[0], parser="ambr"
            )
            du_obj = du_qs[0]
        except Exception as ex:
            logger.debug(
                f"Error trying to retrieve default unit \
            for {mtype_name}"
            )
            logger.exception(ex)

        # appending mapped measurements to parsed worksheet
        for i in range(1, len(mes_data)):
            # dropping records with NaN values
            if not np.isnan(float(mes_data[i])):
                self.worksheet.append(
                    [
                        name,
                        mtype_name,
                        float(mes_data[i]),
                        float(time_data[i]),
                        du_obj.unit.unit_name,
                    ]
                )
