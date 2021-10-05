import collections
import csv
import decimal
import itertools
import logging
import numbers
import re
from dataclasses import dataclass, field
from typing import Dict, List, Sequence, Set, Tuple
from uuid import UUID

import pandas as pd
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from edd.load.models import DefaultUnit, MeasurementNameTransform

from .. import exceptions, reporting

logger = logging.getLogger(__name__)


def _has_any_value(*args):
    # Tests whether any of the provided arguments
    # should be treated as having a valid value
    def is_valid(arg):
        if isinstance(arg, str):
            # assume it's already .strip()ed
            return bool(arg)
        return arg is not None

    return any(filter(is_valid, args))


@dataclass
class ColLayoutDetectionState:
    """
    State of columns built from tabular inputs.

    Captures state associated with detecting a tabular file format while
    scanning the file for rows with one or more required column headers.
    """

    # maps canonical column name -> col index where it was detected
    layout: Dict[str, int] = field(init=False)
    # non-column-header values read from file before any valid column header was found
    non_header_vals: List[str] = field(init=False)
    # maps canonical names of any required columns observed in this row to the list of indexes of
    # columns where they were detected
    obs_req_cols: Dict[str, List[int]] = field(init=False)
    # canonical names of optional columns observed in this row
    obs_opt_cols: Dict[str, List[int]] = field(init=False)

    def __post_init__(self):
        self.layout = {}
        self.non_header_vals = []
        self.obs_req_cols = collections.defaultdict(list)
        self.obs_opt_cols = collections.defaultdict(list)


class TableParser:
    """
    Parser for tabular inputs.

    A parser for the tabular import files that allows supports semi-tolerant
    parsing of user-provided spreadsheets. Basic conditions for parsing are:

    1) Column header labels are fixed or are provided from a short list of
       possibilities, with some reasonable flexibility for capitalization
       and whitespace.
    2) While the labeling is fixed, required columns can be provided in
       any order.
    3) Optional columns will be identified if present.
    4) Any column whose header doesn't match those specified by the file format
       will be ignored (with a warning).
    """

    def __init__(
        self, req_cols: List[str], import_uuid: UUID, opt_cols=None, numeric_cols=None,
    ):
        """
        Initializes the parser using client-provided lists of required and optional column headers.

        :param req_cols: an iterable of human-readable strings containing canonical column
            headers. The parser will take care of tolerating variations in optional leading,
            trailing, and internal whitespace.
        :param opt_cols: an iterable of human-readable strings with canonical names for optional
            column headers. Subject to the same processing as req_cols.
        :param numeric_cols: an  optional list of column names for columns that will be verified
            during parsing to have numeric content.
        """
        super().__init__()
        self.import_uuid: UUID = import_uuid

        # TODO: add support for vector input
        # TODO: add optional column aliases?
        # keep inputs to use as keys and as human-readable for use in err messages
        self.req_cols: List[str] = req_cols
        self.opt_cols = opt_cols if opt_cols is not None else set()
        self.numeric_cols = numeric_cols if numeric_cols is not None else set()

        # maps canonical col name -> observed text (e.g. maybe including whitespace, mixed case,
        # alias)
        self.obs_col_names = {}
        self.obs_col_units = {}

        self._is_excel = None

        # build up regex patterns for each supported column header that allow us to do
        # semi-tolerant matching against user-provided values.
        # maps human-readable col name => pattern
        self.req_col_patterns = self._build_col_header_patterns(req_cols)

        if opt_cols:
            self.opt_col_patterns = self._build_col_header_patterns(self.opt_cols)
        else:
            self.opt_col_patterns = []

        self._ignored_preamble_vals = []

    def _verify_layout(self, header_row_index):
        pass  # children may optionally implement

    def _parse_row(self, cols_list, row_index):
        """
        Parses a data row after columns have been identified.

        Any format errors should be tracked and parsing allowed to continue.

        :param cols_list: columns in the row
        :param row_index: index of the row into the file
        """
        # children must override
        raise NotImplementedError()

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file.

        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped if the content
            was a string
        """
        # children must override
        raise NotImplementedError()

    def _parse_result(self):
        """
        Constructs a ParseResult to capture results of successful parsing.

        Only called after the file has been successfully parsed.
        """
        # children must override
        raise NotImplementedError()

    def _build_col_header_patterns(self, col_headers):
        """
        Builds patterns for parsing of values expected to match during parsing.

        Patterns are constructed to match case-insensitive input and are
        tolerant of leading and trailing whitespace insertions, or of added
        internal whitespace.  Internal whitespace in the input is required, but
        may use different whitespace characters or have additional
        whitespace added.
        """
        reg = r"^\s*{title}\s*$"
        return [
            re.compile(
                reg.format(title=TableParser._process_label(col_header)), re.IGNORECASE
            )
            for col_header in col_headers
        ]

    @staticmethod
    def _process_label(s):
        tokens = re.split(r"\s+", s)  # tokenize label on whitespace
        escaped_tokens = (re.escape(part) for part in tokens)  # escape each part
        return r"\s+".join(escaped_tokens)  # gather together with whitespace regex

    def cell_content_desc(self, content, row_index, col_index):
        """
        Builds a description of the cell content and location.

        :param content: the cell content
        :param row_index: index into the file where this row is located
        :param col_index: index into the file where this column is located
        :return: a string description of the cell content and location
        """
        return f'"{content}" ({self.cell_coords(row_index, col_index)})'

    @staticmethod
    def cell_coords(row_index, col_index):
        """
        Builds a human-readable description of a cell, e.g. "A1"

        :param row_index: 0-indexed row location
        :param col_index: 0-indexed column location
        :return: a standard Excel identifier for the cell, e.g. "A1"
        """
        col_letter = get_column_letter(col_index + 1)
        return f"{col_letter}{row_index + 1}"

    def _parse_rows(self, rows_iter):
        """
        Performs parsing, independent of how the file content is stored.

        :return: the number of non-whitespace rows parsed
        :raise ParseError: if the file format or content was bad
        """
        # Clear out state from any previous use of this parser instance
        self.column_layout = None

        # loop over rows
        for row_index, cols_list in enumerate(rows_iter):

            # identify columns of interest first by looking for required header labels
            if not self.column_layout:
                self.column_layout = self._parse_col_layout(cols_list, row_index)

                if self.column_layout:
                    if self._ignored_preamble_vals:
                        # if ignored values were found# before the required
                        # column headers, report them now. This breaks strict
                        # sequencing of reported warnings, but is probably good
                        # in that ignored columns will be reported first and
                        # emphasized over any ignored preamble
                        reporting.warnings(
                            self.import_uuid,
                            exceptions.IgnoredValueWarning(
                                details=self._ignored_preamble_vals
                            ),
                        )
                    self._verify_layout(row_index)

            # if column labels have been identified, look for data
            else:
                self._parse_row(cols_list, row_index)

        # if we got all the way through the file without finding a single required column header,
        # it's badly formatted
        if not self.column_layout:
            logger.debug(f"Required columns were not read {self.req_cols}")
            # TODO: revert to ", ".join()?
            reporting.raise_errors(
                self.import_uuid, exceptions.RequiredColumnError(details=self.req_cols)
            )

        reporting.raise_errors(self.import_uuid)

        return self._parse_result

    def _parse_col_layout(self, row, row_index):
        """
        Scans a row to see if it contains required column headers for the layout.

        :param row: the row to inspect for column headers
        :param row_index: the index into the file of the row being checked for required col headers
        :return: the column layout if required columns were found, or None otherwise
        """
        state = ColLayoutDetectionState()

        # loop over columns in the current row
        for col_index, cell in enumerate(row):
            cell_content = self._raw_cell_value(cell)

            # skip this cell if it has no non-whitespace content
            # (strings are stripped by _get_raw_value())
            if not cell_content:
                continue

            # ignore non-string cells since they can't be the column headers we're looking for
            if not isinstance(cell_content, str):
                col_desc = self.cell_content_desc(cell_content, row_index, col_index)
                state.non_header_vals.append(col_desc)
                continue

            # check whether column label matches one of the canonical column names
            # specified by the layout
            self._process_col_name(cell_content, row_index, col_index, state)

        # if at least the required columns were found, consider this a successful read
        obs_req_cols_set = set(state.obs_req_cols.keys())
        req_cols_set = set(self.req_cols)
        if obs_req_cols_set == req_cols_set:
            if state.non_header_vals:
                reporting.warnings(
                    self.import_uuid,
                    exceptions.IgnoredColumnWarning(details=state.non_header_vals),
                )
            logger.debug(
                f"Found all {len(self.req_cols)} required column headers in row "
                f"{row_index+1}"
            )
            return state.layout

        # if some--but not all--required columns were found in this row, consider this a
        # poorly-formatted file
        if obs_req_cols_set:
            missing_cols = req_cols_set.difference(state.obs_req_cols.keys())
            # sort the columns alphabetically.  set difference doesn't seem to maintain the initial
            # alphabetical order of required columns during testing
            ordered_missing_cols = sorted(list(missing_cols))
            quoted_names_str = ", ".join([f'"{name}"' for name in ordered_missing_cols])
            msg = _(
                "{found_ct} required columns were found on row {row_num}, but {missing_ct}"
                " others were missing: {missing}"
            ).format(
                found_ct=len(obs_req_cols_set),
                row_num=row_index + 1,
                missing_ct=len(missing_cols),
                missing=quoted_names_str,
            )
            reporting.raise_errors(
                self.import_uuid, exceptions.RequiredColumnError(details=msg)
            )
        else:
            self._ignored_preamble_vals.extend(state.non_header_vals)

        return None

    def obs_col_name(self, canonical_name):
        """
        Gets column label from the file corresponding to the canonical parser name.

        :param canonical_name: the canonical column name.
        :return: the observed column name, or the canonical column name if the column wasn't found
            in the file.
        """
        return self.obs_col_names.get(canonical_name, canonical_name)

    def _parse_and_verify_val(self, val, row_index, col_name):
        """
        Tests a parsed value, and logs an error message if it the observed value doesn't meet
        expectations. Verifies that:

        1) The value is present if required
        2) It's a scalar number, if required
        3) If it's a unit, that it falls within the set of supported units

        :param val: the value read from file
        :param row_index: the index of the file row the value is from
        :param col_name: the canonical name for this column, rather than the one observed in
            the file.
        :return: the value, if it was valid, or None if it failed one ore more validation criteria
        """
        col_index = self.column_layout.get(col_name, None)

        # if value is missing, but required, log an error
        if val is None:
            reporting.add_errors(
                self.import_uuid,
                exceptions.RequiredValueError(
                    subcategory=self.obs_col_name(col_name),
                    details=self.cell_coords(row_index, col_index),
                ),
            )
            return val

        # if observed value is a string,
        if isinstance(val, str):
            if not val:
                reporting.add_errors(
                    self.import_uuid,
                    exceptions.RequiredValueError(
                        subcategory=self.obs_col_name(col_name),
                        details=self.cell_coords(row_index, col_index),
                    ),
                )
                return None

            # parse / verify expected numeric values... when read from CSV, they'll have to be
            # parsed
            if col_name in self.numeric_cols:
                return self._parse_num(val, col_name, row_index, col_index)

            return val

        return val

    def _parse_num(self, token, col_name, row_index, col_index):
        try:
            return decimal.Decimal(token)
        except (decimal.InvalidOperation, decimal.Clamped):
            reporting.add_errors(
                self.import_uuid,
                exceptions.InvalidValueError(
                    subcategory=self.obs_col_name(col_name),
                    details=self.cell_content_desc(token, row_index, col_index),
                ),
            )
            return None

    def _parse_col_header(
        self, col_patterns, col_names, row_index, col_index, cell_content
    ):
        """
        Tests the cell content against the canonical column definitions, and extracts the
        canonical column name and optional observed units if it matches.

        :param col_patterns: the iterable of column name patterns to match against
        :param col_names: an iterable of canonical column names, parallel structure to
            col_patterns
        :param row_index: the row index currently being parsed
        :param col_index: the column index currently being parsed
        :param cell_content: the content of the current cell
        """
        # loop over all column patterns, even those already matched.  Otherwise we can't detect
        # duplicates.
        for col_pattern, col_name in zip(col_patterns, col_names):
            if col_pattern.match(cell_content):
                return col_name

        return None

    def _process_col_name(
        self,
        cell_content,
        row_index: int,
        col_index: int,
        state: ColLayoutDetectionState,
    ):
        """
        Process the string content of a cell when searching the file for identifying column headers

        :param cell_content: the string content (non empty, non-whitespace)
        :param state: intermediate state for tracking column layout while it's being detected
        """

        # test whether col name matches the pattern for any required or optional column
        req_name = self._parse_col_header(
            self.req_col_patterns, self.req_cols, row_index, col_index, cell_content
        )
        opt_name = self._parse_col_header(
            self.opt_col_patterns, self.opt_cols, row_index, col_index, cell_content
        )

        # process it according to which pattern it matched
        if req_name:
            logger.debug(
                f'Found required column "{req_name}"'
                f"({self.cell_coords(row_index, col_index)})"
            )

            self._process_col_helper(
                req_name, cell_content, row_index, col_index, state, True
            )
        elif opt_name:
            logger.info(
                f'Found optional column "{opt_name}"'
                f"({self.cell_coords(row_index, col_index)})"
            )
            self._process_col_helper(
                opt_name, cell_content, row_index, col_index, state, False
            )
        else:
            col_desc = self.cell_content_desc(cell_content, row_index, col_index)
            state.non_header_vals.append(col_desc)

    def _process_col_helper(
        self,
        canonical_name: str,
        cell_content,
        row_index: int,
        col_index: int,
        layout_state: ColLayoutDetectionState,
        required_col: bool,
    ):

        obs_cols_dict: Dict[str, List[int]] = (
            layout_state.obs_req_cols if required_col else layout_state.obs_opt_cols
        )

        # if column name is already observed, build a helpful error message
        if canonical_name in obs_cols_dict:
            cols: List[int] = obs_cols_dict[canonical_name]
            subcategory = f'"{canonical_name}"'

            # add entries for both the first and current instance
            details = [
                self.cell_coords(row_index, cols[0]),
                self.cell_coords(row_index, col_index),
            ]
            reporting.add_errors(
                self.import_uuid,
                exceptions.DuplicateColumnError(
                    subcategory=subcategory, details=details
                ),
            )

        # otherwise, save state re: where we found it
        else:
            layout_state.layout[canonical_name] = col_index
            if canonical_name != cell_content:
                self.obs_col_names[canonical_name] = cell_content

        obs_cols_dict[canonical_name].append(col_index)

    def _get_raw_value(self, row, col_name):
        """
        Gets the raw value from the table, regardless of file format.

        :param row: an iterable of columns in this row of the file
        :param col_name: the canonical name of the column whose value should be read
        :return: the value, or None if the col with col_name isn't in the file
        """
        col_index = self.column_layout.get(col_name, None)
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        return self._raw_cell_value(cell)


@dataclass(frozen=True)
class MeasurementParseRecord:
    """
    A record of a single value or set of related values from a payload.

    This object should be flexible enough to capture the full level of detail
    needed to construct a Measurement & MeasurementValue from any file, though
    many formats won't require this level of detail.

    :param loa_name: the name of the line or assay this record applies to.
        Whether it matches to a line or to an assay is determined subsequently
        during the import.
    :param mtype_name: the name of the MeasurementType this record measures.
        e.g. a Uniprot accession ID or PubChem Compound ID
    :param value_format: the Measurement.Format instance that describes layout of the
        data in this record
    :param meta: a dict that maps type_name => value for any metadata
        associated with this record
    :param data: an array of data to store in the database, packed according
        to "value_format"
    :param x_unit_name: the string name for MeasurementUnits on the x-axis in
        this record (e.g. "hours" is built-in)
    :param y_unit_name: the string name for MeasurementUnits for the y-axis in
        this record ( e.g. "g/L" is built in)
    :param src_ids: an iterable of identifiers for file locations this record
        was sourced from. This is important input for constructing helpful
        error messages if problems occur downstream (e.g. in resolving values
        from the file with reference databases)
    """

    loa_name: str
    mtype_name: str
    value_format: str
    data: List[List[numbers.Number]]
    x_unit_name: str
    y_unit_name: str
    # data source(s) within the file for this measurement...for tabular data, an iterable of
    # int row nums or string ranges, e.g. ('1-3', 5, 24). Used to construct helpful / precise
    # error messages
    src_ids: [Tuple[str, ...]]


@dataclass(eq=False)
class ParseResult:
    """
    Standard format for capturing the results of parsing data from a LoadRequest.

    Note this doesn't imply that the import can necessarily succeed, just that
    the baseline of required data in the file has been read and verified for
    the expected basic data type (e.g. numeric, string).
    """

    series_data: Sequence[MeasurementParseRecord]
    # a human-readable identifier for the file portion(s) identified by each
    # MeasurementParseRecord's src_ids. For example, for tabular data, "row" is
    # often used
    record_src: str
    # True if the file contained at least one time value
    any_time: bool
    # True if a time value was found
    # corresponding to each MeasurementParseRecord
    # Since parsed values may be packed differently,
    # it's simpler for the parser to make this determination.
    has_all_times: bool

    # set of unique line or assay names found in the file
    line_or_assay_names: Set[str] = field(init=False)
    # True if every record parsed from file had associated units
    has_all_units: bool = field(init=False)
    # set of unique mtype identifiers (strings) found in the file
    mtypes: Set[str] = field(init=False)
    # set of unique unit names (strings) found in the file or implicit in the format
    units: Set[str] = field(init=False)

    def __post_init__(self):
        # compute unique line / assay names, units, measurement types from the parse records
        line_or_assay_names = set()
        units = set()
        mtypes = set()
        has_all_units = True
        for record in self.series_data:
            line_or_assay_names.add(record.loa_name)
            units.add(record.x_unit_name)
            units.add(record.y_unit_name)
            mtypes.add(record.mtype_name)
            if not record.x_unit_name or not record.y_unit_name:
                has_all_units = False
        # set member fields
        self.line_or_assay_names = frozenset(line_or_assay_names)
        self.has_all_units = has_all_units
        self.mtypes = frozenset(mtypes)
        self.units = frozenset(units)


class ExcelParserMixin:
    def parse(self, file):
        """
        Parses the input as an Excel workbook.

        :param file: a file-like object
        :return: the ParseResult read from file, otherwise None
        :raises OSError: if the file can't be opened
        :raises EDDImportError: if the file format or content is bad
        """
        wb = load_workbook(file, read_only=True, data_only=True)
        logger.debug("In parse(). workbook has %d sheets" % len(wb.worksheets))
        if len(wb.worksheets) > 1:
            sheet_name = wb.sheetnames[0]
            count = len(wb.worksheets) - 1
            message = _(
                'Only the first sheet in your workbook, "{sheet}", was processed. '
                "All other sheets were ignored ({count})."
            ).format(sheet=sheet_name, count=count)
            reporting.warnings(
                self.import_uuid, exceptions.IgnoredWorksheetWarning(details=[message])
            )
        worksheet = wb.worksheets[0]
        return self._parse_rows(worksheet.iter_rows())

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file.

        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped
            if the content was a string
        """
        val = cell.value
        if isinstance(val, str):
            return val.strip().strip("\ufeff")
        return val


class MultiSheetExcelParserMixin:
    def parse(self, file):
        """
        Parses the input as a Pandas dataframe and then
        converts it into an Excel workbook.

        :param file: a file-like object
        :return: the ParseResult read from file, otherwise None
        :raises OSError: if the file can't be opened
        :raises EDDImportError: if the file format or content is bad
        """

        parsed_result = pd.DataFrame()
        wb = pd.read_excel(file, sheet_name=None)

        # for each worksheet in the workbook
        for name, sheet in wb.items():
            parsed_result = self.get_parsed_records(name, sheet, parsed_result)

        # convert parsed_result into a openpyxl worksheet
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(parsed_result, index=True, header=True):
            ws.append(r)

        # passing the rows in the worksheet for verification
        # and processing in database
        return self._parse_rows(ws.iter_rows())

    def get_parsed_records(self, name, sheet, parsed_result):

        # for every two columns in the worksheet
        # corresponding to each measurement type in the sheet
        mapper = MeasurementMapper()
        for i in range(0, int(sheet.shape[1]), 2):
            two_cols = sheet[sheet.columns[i : i + 2]]

            # dropping all rows with nan values in the worksheet
            two_cols = two_cols.dropna()
            if not two_cols.dropna().empty:
                # decimate the data
                two_cols = two_cols.iloc[::10, :]

                # using mapper to map data into the EDD import format
                # and convert in to a pandas dataframe
                # set the line name and the dataframe with the two columns
                # with data for the next measurement type
                mapper.set_line_name(name)
                mapper.set_data(two_cols)
                parsed_df = mapper.map_data()
                parsed_result = parsed_result.append(parsed_df)
        return parsed_result

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file.

        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped
            if the content was a string
        """
        val = cell.value
        if isinstance(val, str):
            return val.strip()
        return val


class MeasurementMapper:

    loa_name: str
    df: pd.DataFrame
    units: Dict
    mes_unit_map: Dict
    mtype_map = Dict

    def __init__(self, name=None, df=None):
        self.loa_name = name
        self.df = df

        # get the default units for measurements from the database
        self.mes_unit_map = {}
        du_obj = DefaultUnit.objects.all()
        for obj in du_obj:
            self.mes_unit_map[obj.measurement_type.type_name] = obj.unit.unit_name

        # get the mapping between input measurement type names and expected
        # edd measurement type names from database
        self.mtype_map = {}
        mtype_obj = MeasurementNameTransform.objects.all()
        for obj in mtype_obj:
            self.mtype_map[obj.input_type_name] = obj.edd_type_name

    def set_line_name(self, name):
        self.loa_name = name

    def set_data(self, df):
        self.df = df

    def map_data(self):

        mtype_name = self.df[self.df.columns[1:2]].columns.values[0]
        self.df["Line Name"] = self.loa_name
        self.df.columns.values[0] = "Time"
        self.df.columns.values[1] = "Value"

        # check measurement type to rename for EDD
        if mtype_name in self.mtype_map.keys():
            self.df["Measurement Type"] = self.mtype_map[mtype_name]
            mtype_name = self.mtype_map[mtype_name]
        else:
            self.df["Measurement Type"] = mtype_name

        self.df["Units"] = self.mes_unit_map[mtype_name]
        # dropping records with NaN values
        self.df = self.df[self.df["Value"].notna()]

        return self.df


class CsvParserMixin:
    def parse(self, file):
        """
        Parses the input as CSV.

        :param file: a file-like object
        :return: the ParseResult read from file, otherwise None
        :raises OSError: if the file can't be opened
        :raises EDDImportError: if the file format or content is bad
        """
        reader = csv.reader(file)
        return self._parse_rows(reader)

    def _raw_cell_value(self, cell):
        return cell.strip().strip("\ufeff")


def build_src_summary(sources, convert_ints=False):
    """
    Condenses sources by merging consecutive integers into a string range.

    For example, [1, 2, 3, 5] => ["1-3", 5]. Non-integers are simply copied to the result,
    so for example [1, 2, 3, "18-36", 42] =>  ["1-3", "18-36", 42].

    :param sources: an iterable of integer values, e.g. line numbers
    :param convert_ints: if True, convert any single integers in the result to strings
    :return: a list of source ranges, where ranges are strings and any single sources are
        either ints or strings as determined by convert_ints
    """
    src_ranges = []

    def pairwise(iterable):
        # make two independent iterators
        a, b = itertools.tee(iterable)
        # advance the second by one
        next(b, None)
        # return pairwise groupings from two iterators
        return zip(a, b)

    range_start = None
    next_val = None
    if len(sources) < 2:
        return [str(v) if convert_ints else v for v in sources]
    for current, next_val in pairwise(sources):
        if range_start is None and isinstance(current, int):
            range_start = current
        if not isinstance(current, int):
            src_ranges.append(current)
        elif next_val != current + 1:
            if range_start != current:
                src_ranges.append(f"{range_start}-{current}")
            else:
                src_ranges.append(str(current) if convert_ints else current)
            range_start = None
    else:
        if range_start is not None:
            src_ranges.append(f"{range_start}-{next_val}")
        else:
            src_ranges.append(str(next_val) if convert_ints else next_val)
    return src_ranges
