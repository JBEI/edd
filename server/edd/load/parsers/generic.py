from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from main.models import Measurement

from ..reporting import raise_errors
from .core import (
    CsvParserMixin,
    ExcelParserMixin,
    MeasurementParseRecord,
    ParseResult,
    TableParser,
    _has_any_value,
)


class GenericImportParser(TableParser):
    """
    Parser for EDD's "Generic" import file: a normalized, simple, accessible,
    and machine-readable tabular format designed largely for automated data import.
    """

    default_columns = ["Line Name", "Measurement Type", "Value", "Time", "Units"]

    def __init__(self, import_uuid: UUID):
        super().__init__(
            req_cols=self.default_columns,
            import_uuid=import_uuid,
            numeric_cols=["Time", "Value"],
        )
        self._measurements = []

    def _verify_layout(self, header_row_index):
        pass

    def _parse_row(self, cells_list, row_index):
        # extract raw values from in-use cols in this row
        # Note: "Line Name" header is used for simplicity for most users, even though it may
        # actually match assay name during some imports....the difference should often be
        # transparent.
        loa_name = self._get_raw_value(cells_list, "Line Name")
        mtype = self._get_raw_value(cells_list, "Measurement Type")
        val = self._get_raw_value(cells_list, "Value")
        time = self._get_raw_value(cells_list, "Time")
        units = self._get_raw_value(cells_list, "Units")

        # skip the row entirely if no in-use column has a value in it
        any_value = _has_any_value(loa_name, mtype, val, time, units)
        if not any_value:
            return None

        # now that we've seen at least a single value in the row, do more rigorous parsing /
        # verification of the values
        loa_name = self._parse_and_verify_val(loa_name, row_index, "Line Name")
        mtype = self._parse_and_verify_val(mtype, row_index, "Measurement Type")
        val = self._parse_and_verify_val(val, row_index, "Value")
        time = self._parse_and_verify_val(time, row_index, "Time")
        units = self._parse_and_verify_val(units, row_index, "Units")

        # store x and y as arrays to match underlying field types in MeasurementValue...
        # easier comparisons along the way with existing study data
        data = [[time], [val]]

        # always convert loa_name to a string, since Excel / openpyxl may convert numeric names to
        # numbers, which won't match when compared to the study
        loa_name = str(loa_name)
        m = MeasurementParseRecord(
            loa_name=loa_name,
            mtype_name=mtype,
            # TODO: revisit value_format when adding vector support
            value_format=Measurement.Format.SCALAR,
            data=data,
            x_unit_name="hours",  # assumed hours for time, included in EDD's bootstrap.json
            y_unit_name=units,
            src_ids=(row_index + 1,),
        )
        self._measurements.append(m)

    @property
    def _parse_result(self):
        raise_errors(self.import_uuid)

        # note successful parsing of this format implies that every record has a time, since
        # time is a required column for the format
        return ParseResult(
            series_data=self._measurements,
            any_time=True,  # required by format
            has_all_times=True,  # required by format
            record_src="row",
        )


class GenericExcelParser(ExcelParserMixin, GenericImportParser):
    pass


class GenericCsvParser(CsvParserMixin, GenericImportParser):
    pass


class AmbrExcelParser(ExcelParserMixin, CsvParserMixin):
    def parse(self):
        """
        Overriding the parse function for ExcelParserMixin to enable processing multiple sheets
        for Ambr250 data.
        """
        """
        Parses the input as an Excel workbook.

        :param file: a file-like object
        :return: the ParseResult read from file, otherwise None
        :raises OSError: if the file can't be opened
        :raises EDDImportError: if the file format or content is bad
        """

        # reading the excel sheet and decimating it by selecting every 10th record
        sheets_dict = pd.read_excel(
            file, sheet_name=None, skiprows=lambda x: x % 10 > 0
        )

        # for each sheet in the excel sheet if the sheet has multiple sheets
        for bioreactor_name, sheet in sheets_dict.items():

            # return self._parse_rows(worksheet.iter_rows())
            # for each individual measurement type taken in the
            # ["Line Name", "Measurement Type", "Value", "Time", "Units"] format
            for measurement_worksheet in self.process_ambr_data(bioreactor_name, sheet):
                self._parse_rows(measurement_worksheet.iter_rows())

    def process_ambr_data(self, bioreactor_name, sheet):

        units = {
            "Temperature": "°C",
            "Stir speed": "rpm",
            "pH": "n/a",
            "Air flow": "lpm",
            "DO": "% maximum measured",
            "Volume": "mL",
            "OUR": "mM/L/h",
            "CER": "mM/L/h",
            "RQ": "n/a",
            "Feed#1 volume pumped": "mL",
            "Antifoam volume pumped": "mL",
            "Acid volume pumped": "mL",
            "Base volume pumped": "mL",
            "Volume - sampled": "mL",
        }

        # for bioreactor_name, sheet in sheets_dict.items():
        second_ind = 2
        # Iterate through every pair of columns for each measurement_type
        while second_ind <= len(sheet.columns) + 1:
            df = sheet.iloc[:, second_ind - 2 : second_ind]
            timestamps = df.iloc[:, 0].name
            line_name = bioreactor_name
            df = df.dropna(subset=[timestamps]).fillna(0)

            reformatted_data = {}

            # Catch "Volume of inocula" column (not included in google doc)
            if df.columns[1] in units:
                unit = units[df.columns[1]]
            else:
                second_ind += 2
                continue

            # Hardcoded measurement type renaming for certain columns
            measurement_type = df.columns[1]
            if measurement_type == "Volume - sampled":
                measurement_type = "Volume sampled"
            elif measurement_type == "Feed#1 volume pumped":
                measurement_type = "Feed volume pumped"
            elif measurement_type == "Temperature":
                measurement_type = "Vessel temperature"
            elif measurement_type == "Volume":
                measurement_type = "Working volume"

            reformatted_data["Line Name"] = [line_name for _ in range(len(df.index))]
            reformatted_data["Measurement Type"] = [
                measurement_type for _ in range(len(df.index))
            ]
            reformatted_data["Time"] = df.iloc[:, 0]
            reformatted_data["Units"] = [unit for _ in range(len(df.index))]

            # Convert "Air flow" data from mL/min to lpm
            if measurement_type == "Air flow":
                reformatted_data["Value"] = df.iloc[:, 1].div(1000)
            else:
                reformatted_data["Value"] = df.iloc[:, 1]

            order = ["Line Name", "Measurement Type", "Time", "Value", "Units"]
            reformatted_df = pd.DataFrame(data=reformatted_data)[order]

            # convert pandas dataframe to openpyxl worksheet
            wb = Workbook()
            ws = wb.active
            for r in dataframe_to_rows(reformatted_df, index=True, header=True):
                ws.append(r)

            second_ind += 2
            
            # yeilding each measurement type as a work sheet to be processed in the parse function
            yield ws