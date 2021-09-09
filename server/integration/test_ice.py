"""Integration tests for ICE."""

import itertools
from io import BytesIO
from unittest.mock import patch

from django.test import override_settings, tag
from django.urls import reverse
from faker import Faker
from openpyxl.workbook import Workbook
from requests import codes

from edd import TestCase
from edd.load.resolver import TypeResolver
from edd.load.tests import factory as load_factory
from edd.profile.factory import UserFactory
from edd.search.registry import AdminRegistry, RegistryError, StrainRegistry
from main import models
from main.tests import factory

faker = Faker()


@tag("integration")
class IceIntegrationTests(TestCase):
    """Sets of tests to validate communication between EDD and ICE."""

    def setUp(self):
        self.registry = StrainRegistry()

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        registry = AdminRegistry()
        with registry.login():
            try:
                # make sure ICE has users matching EDD users
                cls._ensureTestUsers(registry)
                # populate ICE with some strains
                entries = cls._populateTestStrains(registry)
                # add strains to a folder
                cls._populateTestFolder(registry, entries)
            except Exception as e:
                cls.tearDownClass()
                raise e

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.admin_ice_user = UserFactory(email="admin@example.org")
        cls.read_ice_user = UserFactory(email="reader@example.org")
        cls.none_ice_user = UserFactory(email="none@example.org")

    @classmethod
    def _ensureTestUsers(cls, registry):
        # make an admin users
        admin_id = registry.get_user_id(cls.admin_ice_user)
        if admin_id is None:
            admin_id = registry.create_admin(cls.admin_ice_user)
        # stash the ICE user ID for later use in permissions
        cls.admin_ice_user._ice_id = admin_id
        # make regular users
        for user in [cls.read_ice_user, cls.none_ice_user]:
            user_id = registry.get_user_id(user)
            if user_id is None:
                user_id = registry.create_user(user)
            # stash the ICE user ID for later use in permissions
            user._ice_id = user_id

    @classmethod
    def _populateTestStrains(cls, registry):
        with factory.load_test_file("ice_entries.csv") as entries_file:
            registry.bulk_upload(entries_file)
        # fetch the part IDs
        it = registry.iter_entries(sort="created", asc="false", limit=10)
        entries = list(itertools.islice(it, 10))
        # set read permissions on some of the created strains
        for entry in entries[:5]:
            entry.set_permission(cls.read_ice_user._ice_id)
        # stash the part IDs for use in tests
        cls.entry_ids = [entry.part_id for entry in entries]
        return entries

    @classmethod
    def _populateTestFolder(cls, registry, entries):
        # create folder
        cls.folder = registry.create_folder(faker.catch_phrase())
        # set it to public
        # ice.session.put(ice_url(f"/folders/{cls.folder_id}/permissions/public"))
        # add our parts to it
        cls.folder.add_entries(entries)

    @override_settings(ICE_URL=None)
    def test_no_configured_ICE_raises_error(self):
        with self.assertRaises(RegistryError):
            self.registry.base_url

    @override_settings(ICE_URL="https://example.com")
    def test_configured_ICE_without_trailing_slash(self):
        assert self.registry.base_url == "https://example.com"

    @override_settings(ICE_URL="https://example.com/")
    def test_configured_ICE_strips_trailing_slash(self):
        assert self.registry.base_url == "https://example.com"

    def test_action_without_login_throws_error(self):
        with self.assertRaises(RegistryError):
            self.registry.get_entry("42")

    def test_action_after_logout_throws_error(self):
        with self.registry.login(self.none_ice_user), self.assertRaises(RegistryError):
            # force logout
            self.registry.logout()
            # this should trigger RegistryError following logout
            self.registry.get_entry("42")

    def test_admin_find_parts(self):
        with self.registry.login(self.admin_ice_user):
            # verify that admin user finds all parts
            entries = self.registry.list_entries(sort="created", asc="false")
            ids = {entry.part_id for entry in entries}
            assert ids.issuperset(self.entry_ids)

    def test_admin_read_part(self):
        with self.registry.login(self.admin_ice_user):
            # verify that admin user can load specific entry
            entry = self.registry.get_entry(self.entry_ids[0])
            assert entry is not None
            entry = self.registry.get_entry(self.entry_ids[5])
            assert entry is not None

    def test_reader_find_parts(self):
        with self.registry.login(self.read_ice_user):
            # verify that reader user finds the five parts with permissions set
            entries = self.registry.list_entries(sort="created", asc="false")
            ids = {entry.part_id for entry in entries}
            readable = set(self.entry_ids[:5])
            unreadable = set(self.entry_ids[5:])
            assert ids.issuperset(readable)
            assert len(ids.intersection(unreadable)) == 0

    def test_reader_read_part(self):
        with self.registry.login(self.read_ice_user):
            # verify that reader user can load specific entry with permission
            entry = self.registry.get_entry(self.entry_ids[0])
            assert entry is not None

    def test_reader_cannot_read_part_without_permission(self):
        with self.registry.login(self.read_ice_user):
            # verify that reader user cannot load entry without permission
            with self.assertRaises(RegistryError):
                self.registry.get_entry(self.entry_ids[5])

    def test_none_find_parts(self):
        with self.registry.login(self.none_ice_user):
            # verify that user with no permissions finds no parts
            entries = self.registry.list_entries(sort="created", asc="false")
            ids = {entry.part_id for entry in entries}
            assert len(ids.intersection(self.entry_ids)) == 0

    def test_none_read_part(self):
        with self.registry.login(self.none_ice_user):
            # verify no permission user cannot load entries
            with self.assertRaises(RegistryError):
                self.registry.get_entry(self.entry_ids[0])
            with self.assertRaises(RegistryError):
                self.registry.get_entry(self.entry_ids[5])

    def test_upload_links_admin(self):
        admin_study = factory.StudyFactory()
        admin_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.admin_ice_user
        )
        response = self._run_upload(self.entry_ids, admin_study, self.admin_ice_user)
        # should return OK from upload
        self.assertEqual(response.status_code, codes.ok, response.content)
        # there should be 10 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=admin_study).distinct().count(), 10
        )
        # TODO: cannot check links because tests in transaction that ultimately calls rollback()
        # Celery task is only ever submitted when the transaction calls commit() successfully

    def test_upload_links_reader(self):
        reader_study = factory.StudyFactory()
        reader_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.read_ice_user
        )
        # should return 500 error on uploading admin-only strains
        # skip testing this until ICE-90 is resolved
        # response = self._run_upload(self.entry_ids, reader_study, self.read_ice_user)
        # self.assertEqual(response.status_code, codes.server_error)
        # should return OK on uploading readable strains
        response = self._run_upload(
            self.entry_ids[:5], reader_study, self.read_ice_user
        )
        self.assertEqual(response.status_code, codes.ok, response.content)
        # there should be 5 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=reader_study).distinct().count(), 5
        )
        # TODO: cannot check links because tests in transaction that ultimately calls rollback()
        # Celery task is only ever submitted when the transaction calls commit() successfully

    def test_upload_links_none(self):
        none_study = factory.StudyFactory()
        none_study.userpermission_set.update_or_create(
            permission_type=models.StudyPermission.WRITE, user=self.none_ice_user
        )
        # should return 400 error on uploading admin-only strains
        response = self._run_upload(self.entry_ids, none_study, self.none_ice_user)
        self.assertEqual(response.status_code, codes.bad_request, response.content)
        # should return 400 error on uploading reader-only strains
        response = self._run_upload(self.entry_ids[:5], none_study, self.none_ice_user)
        self.assertEqual(response.status_code, codes.bad_request)
        # there should be 0 strains on the study
        self.assertEqual(
            models.Strain.objects.filter(line__study=none_study).distinct().count(), 0
        )

    def test_get_folder_known_id_admin_user(self):
        with self.registry.login(self.admin_ice_user):
            folder = self.registry.get_folder(self.folder.folder_id)
            assert folder is not None
            assert folder.folder_id == self.folder.folder_id
            assert folder.name == self.folder.name

    def test_get_folder_known_id_reader_user(self):
        with self.registry.login(self.read_ice_user):
            with self.assertRaises(RegistryError):
                self.registry.get_folder(self.folder.folder_id)

    def test_get_folder_known_id_none_user(self):
        with self.registry.login(self.none_ice_user):
            with self.assertRaises(RegistryError):
                self.registry.get_folder(self.folder.folder_id)

    def test_get_folder_known_bad_id(self):
        with self.registry.login(self.admin_ice_user):
            with self.assertRaises(RegistryError):
                self.registry.get_folder(self.folder.folder_id + 1)

    def test_get_folder_entries(self):
        with self.registry.login(self.admin_ice_user):
            folder = self.registry.get_folder(self.folder.folder_id)
            entries = folder.list_entries()
            assert len(list(entries)) == 10

    def test_list_folders(self):
        with self.registry.login(self.none_ice_user):
            # we're not setting up any FEATURED folders
            folders = self.registry.list_folders("FEATURED")
            assert len(folders) == 0

    def test_create_folder_upstream_failure(self):
        with self.registry.login(
            self.none_ice_user
        ), self._request_failure(), self.assertRaises(RegistryError):
            self.registry.create_folder("Special Folder")

    def test_iter_entries_upstream_failure(self):
        with self.registry.login(
            self.none_ice_user
        ), self._request_failure(), self.assertRaises(RegistryError):
            # this is a generator, so must force iterate it to get Exception
            next(self.registry.iter_entries())

    def test_list_entries_upstream_failure(self):
        with self.registry.login(
            self.none_ice_user
        ), self._request_failure(), self.assertRaises(RegistryError):
            self.registry.list_entries()

    def test_list_folders_upstream_failure(self):
        with self.registry.login(
            self.none_ice_user
        ), self._request_failure(), self.assertRaises(RegistryError):
            self.registry.list_folders()

    def test_search(self):
        with self.registry.login(self.admin_ice_user):
            # pRS426 is one of the items in the ice_entries.csv file
            results = self.registry.search("pRS426")
            # multiple matching entries, check that one is found
            next(results)

    def test_search_upstream_failure(self):
        with self.registry.login(
            self.none_ice_user
        ), self._request_failure(), self.assertRaises(RegistryError):
            results = self.registry.search("pRS426")
            # iterating results throws RegistryError
            next(results)

    def test_yield_paged_records_will_terminate(self):
        # cover the branch where _yield_paged_records breaks after getting empty results
        items = self.registry._yield_paged_records(lambda count: [])
        assert len(list(items)) == 0

    def test_set_permission_as_normal_user(self):
        with self.registry.login(self.read_ice_user), self.assertRaises(RegistryError):
            entry = self.registry.get_entry(self.entry_ids[0])
            entry.set_permission(self.none_ice_user._ice_id)

    def test_ice_admin_create_admin(self):
        user = UserFactory()
        admin_ice = AdminRegistry()
        with admin_ice.login():
            created_id = admin_ice.create_admin(user)
            found_id = admin_ice.get_user_id(user)
            assert created_id == found_id

    def test_ice_admin_create_user(self):
        user = UserFactory()
        admin_ice = AdminRegistry()
        with admin_ice.login():
            created_id = admin_ice.create_user(user)
            found_id = admin_ice.get_user_id(user)
            assert created_id == found_id

    def test_ice_admin_unknown_user(self):
        user = UserFactory()
        admin_ice = AdminRegistry()
        with admin_ice.login():
            found_id = admin_ice.get_user_id(user)
            assert found_id is None

    def test_ice_admin_get_user_upstream_failure(self):
        admin_ice = AdminRegistry()
        with admin_ice.login():
            admin_failure = patch.object(
                admin_ice.session, "send", side_effect=ValueError("Dummy Exception"),
            )
            with admin_failure, self.assertRaises(RegistryError):
                admin_ice.get_user_id(self.admin_ice_user)

    def test_ice_admin_create_user_twice(self):
        admin_ice = AdminRegistry()
        with admin_ice.login(), self.assertRaises(RegistryError):
            admin_ice.create_user(self.admin_ice_user)

    def test_ice_admin_bulk_upload_upstream_failure(self):
        admin_ice = AdminRegistry()
        with admin_ice.login():
            admin_failure = patch.object(
                admin_ice.session, "send", side_effect=ValueError("Dummy Exception"),
            )
            with admin_failure, self.assertRaises(RegistryError):
                admin_ice.bulk_upload(BytesIO(b""))

    def test_add_remove_experiment_link(self):
        with self.registry.login(self.admin_ice_user):
            study = factory.StudyFactory()
            study_url = f"https://edd.example.org/s/{study.slug}/"
            entry = self.registry.get_entry(self.entry_ids[3])
            entry.add_link(study.name, study_url)
            # verify link exists on a given entry
            found_links = list(entry.list_links())
            assert len(found_links) == 1
            # verify that removal works
            for link in found_links:
                entry.remove_link(link[0])
            found_links = list(entry.list_links())
            assert len(found_links) == 0

    def test_link_upstream_failures(self):
        with self.registry.login(self.admin_ice_user):
            entry = self.registry.get_entry(self.entry_ids[0])
            with self._request_failure(), self.assertRaises(RegistryError):
                entry.add_link("label", "url")
            with self._request_failure(), self.assertRaises(RegistryError):
                entry.remove_link("42")
            with self._request_failure(), self.assertRaises(RegistryError):
                # iterating results to throw RegistryError
                next(entry.list_links())

    def test_folder_add_list_entries_upstream_failure(self):
        with self.registry.login(self.admin_ice_user):
            folder = self.registry.get_folder(self.folder.folder_id)
            with self._request_failure(), self.assertRaises(RegistryError):
                folder.add_entries([])
            with self._request_failure(), self.assertRaises(RegistryError):
                folder.list_entries()

    def test_ice_protein_link(self):
        category = load_factory.CategoryFactory.build(
            type_group=models.MeasurementType.Group.PROTEINID,
        )
        resolver = TypeResolver(self.read_ice_user, category)
        protein = resolver.lookup_type(self.entry_ids[0])
        assert isinstance(protein, models.ProteinIdentifier)
        assert protein.accession_id == self.entry_ids[0]

    def test_ice_protein_link_twice(self):
        category = load_factory.CategoryFactory.build(
            type_group=models.MeasurementType.Group.PROTEINID,
        )
        resolver = TypeResolver(self.read_ice_user, category)
        protein = resolver.lookup_type(self.entry_ids[0])
        # running lookup_type again does not throw ValidationError
        resolver.lookup_type(self.entry_ids[0])
        assert isinstance(protein, models.ProteinIdentifier)
        assert protein.accession_id == self.entry_ids[0]

    def _create_workbook(self, parts):
        upload = BytesIO()
        wb = Workbook()
        ws = wb.active
        for i, title in enumerate(["Line Name", "Part ID", "Media"], 1):
            ws.cell(row=1, column=i).value = title
        for i, part in enumerate(parts, 2):
            ws.cell(row=i, column=1).value = part
            ws.cell(row=i, column=2).value = part
            ws.cell(row=i, column=3).value = "M9"
        wb.save(upload)
        upload.name = "description.xlsx"
        upload.content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        upload.seek(0)
        return upload

    def _request_failure(self):
        """
        Use this in a context manager to simulate a failed HTTP request to ICE.
        """
        return patch.object(
            self.registry.session, "send", side_effect=ValueError("Dummy Exception")
        )

    def _run_upload(self, part_ids, study, user):
        upload = self._create_workbook(part_ids)
        self.client.force_login(user)
        response = self.client.post(
            reverse("main:describe:describe", kwargs={"slug": study.slug}),
            data={"file": upload},
        )
        return response
