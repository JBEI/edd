# coding: utf-8

import collections
import csv
import decimal
import logging
import numbers
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, Iterable, List, Sequence, Set, Tuple
from uuid import UUID

from django.utils.translation import ugettext_lazy as _
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from six import string_types

from ..exceptions import (
    DuplicateColumnError,
    EmptyFileError,
    IgnoredColumnWarning,
    IgnoredValueWarning,
    IgnoredWorksheetWarning,
    InvalidValueError,
    RequiredColumnError,
    RequiredValueError,
    add_errors,
    raise_errors,
    warnings,
)

logger = logging.getLogger(__name__)


def _has_any_value(*args):
    """
    Tests whether any of the provided arguments should be treated as having a valid value
    """

    def _valid_value(arg):
        if isinstance(arg, string_types):
            return bool(arg)  # assume it's already .strip()ed
        else:
            return arg is not None

    return any(filter(_valid_value, args))


class BaseFileParser(ABC):
    def __init__(self):
        super().__init__()

    @abstractmethod
    def parse(self, file):
        """
        Parses the file

        :param file: a file-like object (or more likely a Django FieldFile)
        :return: a FileParseResult object containing results of the file parsing operation
        :raises OSError if the file can't be opened or edd_file_importer.exceptions.ParseError
        if the file couldn't be successfully parsed
        """
        pass


@dataclass(init=False)
class ColLayoutDetectionState:
    """
    Captures state associated with detecting a tabular file format while scanning the file for rows
    with one or more required column headers
    """

    # maps canonical column name -> col index where it was detected
    layout: Dict[str, int]

    # non-column-header values read from file before any valid column header was found
    non_header_vals: List[str]

    # maps canonical names of any required columns observed in this row to the list of indexes of
    # columns where they were detected
    obs_req_cols: Dict[str, List[int]]

    # canonical names of optional columns observed in this row
    obs_opt_cols: Dict[str, List[int]]

    def __init__(self):
        # the motivation for having a custom __init__() -- use defaultdicts.
        # note that __setattr__ is the only way to mutate data members of a dataclass
        object.__setattr__(self, "obs_req_cols", collections.defaultdict(list))
        object.__setattr__(self, "obs_opt_cols", collections.defaultdict(list))

        object.__setattr__(self, "layout", {})
        object.__setattr__(self, "non_header_vals", [])


class TableParser(BaseFileParser, ABC):
    """
    A parser for the tabular import files that allows supports semi-tolerant parsing of
    user-provided spreadsheets. Basic conditions for parsing are:

    1) Column header labels are fixed or are provided from a short list of possibilities, with some
    reasonable flexibility for capitalization & whitespace.
    2) While the labeling is fixed, required columns can be provided in any order
    3) Optional columns will be identified if present
    4) Any column whose header doesn't match those specified by the file format will be
        ignored (with a warning)
    """

    # TODO: capture common method parameters as a class to simplify interface
    def __init__(
        self,
        req_cols: List[str],
        import_uuid: UUID,
        opt_cols=None,
        value_opt_cols=None,
        numeric_cols=None,
        supported_units=None,
    ):
        """
        Initializes the parser using client-provided lists of required and optional column headers.

        :param req_cols: an iterable of human-readable strings containing canonical column
            headers. The parser will take care of tolerating variations in optional leading,
            trailing, and internal whitespace.
        :param opt_cols: an iterable of human-readable strings with canonical names for optional
            column headers. Subject to the same processing as req_cols.
        :param value_opt_cols: an optional iterable of column headers from either or both of
            req_cols & opt_cols. Any col header in value_opt_cols will have missing
            values tolerated in any row.  Any column headers not listed here are by default
            required to have a value in every row, assuming the column is present.
        :param numeric_cols: an  optional list of column names for columns that will be verified
            during parsing to have numeric content.
        """
        super().__init__()
        self.import_uuid: UUID = import_uuid

        # TODO: add support for vector input
        # TODO: add optional column aliases?
        # keep inputs to use as keys and as human-readable for use in err messages
        self.req_cols: List[str] = req_cols
        self.opt_cols = opt_cols if opt_cols is not None else set()
        self.value_opt_cols = value_opt_cols if value_opt_cols is not None else set()
        self.numeric_cols = numeric_cols if numeric_cols is not None else set()
        self.supported_units = supported_units if supported_units is not None else {}

        # maps canonical col name -> observed text (e.g. maybe including whitespace, mixed case,
        # alias)
        self.obs_col_names = {}
        self.obs_col_units = {}

        self._is_excel = None

        # build up regex patterns for each supported column header that allow us to do
        # semi-tolerant matching against user-provided values.
        # maps human-readable col name => pattern
        self.req_col_patterns = self._build_col_header_patterns(req_cols)

        if opt_cols:
            self.opt_col_patterns = self._build_col_header_patterns(self.opt_cols)
        else:
            self.opt_col_patterns = []

        self._build_unit_patterns()

        self._ignored_preamble_vals = []

    def _verify_layout(self, header_row_index):
        pass  # children may optionally implement

    def _parse_row(self, cols_list, row_index):
        """
        Parses a data row after columns have been identified. Any format errors should be tracked
        and parsing allowed to continue

        :param cols_list: columns in the row
        :param row_index: index of the row into the file
        """
        raise NotImplementedError()  # children must implement

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file

        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped if the content
        was a string
        """
        # children must override
        raise NotImplementedError()

    def _parse_result(self):
        """
        Constructs a FileParseResult to capture results of successful parsing. Only called after
        the file has been successfully parsed.
        """
        # children must override
        raise NotImplementedError()

    def _build_col_header_patterns(self, col_headers):
        """
        Builds patterns to facilitate reasonably tolerant parsing of fixed string values expected
        to match during parsing.

        Patterns are constructed to match case-insensitive input and are tolerant of leading and
        trailing whitespace insertions, or of added internal whitespace.  Internal whitespace in
        the input is required, but may use different whitespace characters or have additional
        whitespace added.
        """
        reg = r"^\s*{title}\s*$"
        return [
            re.compile(
                reg.format(title=TableParser._process_label(col_header)), re.IGNORECASE
            )
            for col_header in col_headers
        ]

    def _build_unit_patterns(self):
        self._unit_patterns = {}
        for col, units in self.supported_units.items():
            logger.debug(f'Building unit patterns for column "{col}": {units}')
            # note : maintaining case-sensitivity is important! SI units use case!
            vals = "|".join([TableParser._process_label(unit) for unit in units])
            pat = re.compile(r"^\s*({vals})\s*$".format(vals=vals))
            self._unit_patterns[col] = pat

    @staticmethod
    def _process_label(s):
        tokens = re.split(r"\s+", s)  # tokenize label on whitespace
        escaped_tokens = (re.escape(part) for part in tokens)  # escape each part
        return r"\s+".join(escaped_tokens)  # gather together with whitespace regex

    def cell_content_desc(self, content, row_index, col_index):
        """
        Builds a succinct, human-readable description of the cell content and location for use in
        building helpful error messages.

        :param content: the cell content
        :param row_index: index into the file where this row is located
        :param col_index: index into the file where this column is located
        :return: a string description of the cell content and location
        """
        return f'"{content}" ({self.cell_coords(row_index, col_index)})'

    @staticmethod
    def cell_coords(row_index, col_index):
        """
        Builds a human-readable description of a cell, e.g. "A1"

        :param row_index: 0-indexed row location
        :param col_index: 0-indexed column location
        :return: a standard Excel identifier for the cell, e.g. "A1"
        """
        col_letter = get_column_letter(col_index + 1)
        return f"{col_letter}{row_index + 1}"

    def _parse_rows(self, rows_iter):
        """
        A workhorse method that performs parsing, independent of how the file content is stored
        (CSV or Excel).
        :return: the number of non-whitespace rows parsed
        :raise ParseError: if the file format or content was bad
        """
        # Clear out state from any previous use of this parser instance
        self.column_layout = None

        # loop over rows
        for row_index, cols_list in enumerate(rows_iter):

            # identify columns of interest first by looking for required header labels
            if not self.column_layout:
                self.column_layout = self._parse_col_layout(cols_list, row_index)

                if self.column_layout:
                    if self._ignored_preamble_vals:
                        # if ignored values were found before the required column headers,
                        # report them now.  This breaks strict sequencing of reported warnings,
                        # but is probably good in that ignored columns will be reported first
                        # and emphasized over any ignored preamble
                        warnings(
                            self.import_uuid,
                            IgnoredValueWarning(details=self._ignored_preamble_vals),
                        )
                    self._verify_layout(row_index)

            # if column labels have been identified, look for data
            else:
                self._parse_row(cols_list, row_index)

        # if we got all the way through the file without finding a single required column header,
        # it's badly formatted
        if not self.column_layout:
            logger.debug(f"Required columns were not read {self.req_cols}")
            # TODO: revert to ", ".join()?
            raise_errors(self.import_uuid, RequiredColumnError(details=self.req_cols))

        raise_errors(self.import_uuid)

        return self._parse_result

    def _parse_col_layout(self, row, row_index):
        """
        Scans a row of the file to see if it contains required column headers that define the file
        layout.

        :param row: the row to inspect for column headers
        :param row_index: the index into the file of the row being checked for required col headers
        :return: the column layout if required columns were found, or None otherwise
        """
        state = ColLayoutDetectionState()

        ###########################################################################################
        # loop over columns in the current row
        ###########################################################################################
        for col_index, cell in enumerate(row):
            cell_content = self._raw_cell_value(cell)

            # skip this cell if it has no non-whitespace content
            # (strings are stripped by _get_raw_value())
            if not cell_content:
                continue

            # ignore non-string cells since they can't be the column headers we're looking for
            if not isinstance(cell_content, string_types):
                if cell_content is not None:
                    col_desc = self.cell_content_desc(
                        cell_content, row_index, col_index
                    )
                    state.non_header_vals.append(col_desc)
                continue

            #######################################################################################
            # check whether column label matches one of the canonical column names specified by
            # the format
            #######################################################################################
            self._process_col_name(cell_content, row_index, col_index, state)

        # if at least the required columns were found, consider this a successful read
        obs_req_cols_set = set(state.obs_req_cols.keys())
        req_cols_set = set(self.req_cols)
        if obs_req_cols_set == req_cols_set:
            if state.non_header_vals:
                warnings(
                    self.import_uuid,
                    IgnoredColumnWarning(details=state.non_header_vals),
                )
            logger.debug(
                f"Found all {len(self.req_cols)} required column headers in row "
                f"{row_index+1}"
            )
            return state.layout

        # if some--but not all--required columns were found in this row, consider this a
        # poorly-formatted file
        if obs_req_cols_set:
            missing_cols = req_cols_set.difference(state.obs_req_cols.keys())
            # sort the columns alphabetically.  set difference doesn't seem to maintain the initial
            # alphabetical order of required columns during testing
            ordered_missing_cols = sorted(list(missing_cols))
            quoted_names_str = ", ".join(
                ['"{name}"'.format(name=name) for name in ordered_missing_cols]
            )
            msg = _(
                "{found_ct} required columns were found on row {row_num}, but {missing_ct}"
                " others were missing: {missing}"
            ).format(
                found_ct=len(obs_req_cols_set),
                row_num=row_index + 1,
                missing_ct=len(missing_cols),
                missing=quoted_names_str,
            )
            logger.debug(msg)
            raise_errors(self.import_uuid, RequiredColumnError(details=msg))
        else:
            self._ignored_preamble_vals.extend(state.non_header_vals)

        return None

    def obs_col_name(self, canonical_name):
        """
        Gets the observed column label from the file corresponding to the canonical name understood
        by the parser.

        :param canonical_name: the canonical column name.
        :return: the observed column name, or the canonical column name if the column wasn't found
            in the file.
        """
        return self.obs_col_names.get(canonical_name, canonical_name)

    def _parse_and_verify_val(self, val, row_index, col_name):
        """
        Tests a parsed value, and logs an error message if it the observed value doesn't meet
        expectations. Verifies that:

        1) The value is present if required
        2) It's a scalar number, if required
        3) If it's a unit, that it falls within the set of supported units

        :param val: the value read from file
        :param row_index: the index of the file row the value is from
        :param col_name: the canonical name for this column, rather than the one observed in
            the file.
        :return: the value, if it was valid, or None if it failed one ore more validation criteria
        """
        col_index = self.column_layout.get(col_name, None)

        # if parser is asking for a column not found in the file, end early.
        # if it's required, an error will already have been recorded
        if col_index is None:
            logger.warning(f'Column "{col_name}" not found in file')
            return None

        # if value is missing, but required, log an error
        if val is None:
            if col_name not in self.value_opt_cols:
                add_errors(
                    self.import_uuid,
                    RequiredValueError(
                        subcategory=self.obs_col_name(col_name),
                        details=self.cell_coords(row_index, col_index),
                    ),
                )
            return val

        # if observed value is a string,
        if isinstance(val, string_types):
            if (not val) and col_name not in self.value_opt_cols:
                add_errors(
                    self.import_uuid,
                    RequiredValueError(
                        subcategory=self.obs_col_name(col_name),
                        details=self.cell_coords(row_index, col_index),
                    ),
                )
                return None

            # parse / verify expected numeric values... when read from CSV, they'll have to be
            # parsed
            if col_name in self.numeric_cols:
                return self._parse_num(val, col_name, row_index, col_index)

            return val

        # assumption (for now) is that value is numeric. should work for both vector (
        # 1-dimensional) and numeric inputs
        if not isinstance(val, numbers.Number):
            add_errors(
                self.import_uuid,
                InvalidValueError(
                    subcategory=self.obs_col_name(col_name),
                    details=self.cell_coords(row_index, col_index),
                ),
            )

        return val

    def _parse_num(self, token, col_name, row_index, col_index):
        try:
            return decimal.Decimal(token)
        except (decimal.InvalidOperation, decimal.Clamped):
            add_errors(
                self.import_uuid,
                InvalidValueError(
                    subcategory=self.obs_col_name(col_name),
                    details=self.cell_content_desc(token, row_index, col_index),
                ),
            )
            return None

    def _parse_col_header(
        self, col_patterns, col_names, row_index, col_index, cell_content
    ):
        """
        Tests the cell content against the canonical column definitions, and extracts the
        canonical column name and optional observed units if it matches.

        :param col_patterns: the iterable of column name patterns to match against
        :param col_names: an iterable of canonical column names, parallel structure to
            col_patterns
        :param row_index: the row index currently being parsed
        :param col_index: the column index currently being parsed
        :param cell_content: the content of the current cell
        """
        # loop over all column patterns, even those already matched.  Otherwise we can't detect
        # duplicates.
        for col_pattern, col_name in zip(col_patterns, col_names):
            if col_pattern.match(cell_content):
                return col_name

        return None

    def _verify_required_val(self, value, row_index, col_index, col_title):
        if value is None or (isinstance(value, string_types) and not value.strip()):
            add_errors(
                self.import_uuid,
                RequiredValueError(
                    subcategory=col_title,
                    details=self.cell_coords(row_index, col_index),
                ),
            )

    def _process_col_name(
        self,
        cell_content,
        row_index: int,
        col_index: int,
        state: ColLayoutDetectionState,
    ):
        """
        Process the string content of a cell when searching the file for identifying column headers

        :param cell_content: the string content (non empty, non-whitespace)
        :param state: intermediate state for tracking column layout while it's being detected
        """

        # test whether col name matches the pattern for any required or optional column
        req_name = self._parse_col_header(
            self.req_col_patterns, self.req_cols, row_index, col_index, cell_content
        )
        opt_name = self._parse_col_header(
            self.opt_col_patterns, self.opt_cols, row_index, col_index, cell_content
        )

        # process it according to which pattern it matched
        if req_name:
            logger.debug(
                f'Found required column "{req_name}"'
                f"({self.cell_coords(row_index, col_index)})"
            )

            self._process_col_helper(
                req_name, cell_content, row_index, col_index, state, True
            )
        elif opt_name:
            logger.info(
                f'Found optional column "{opt_name}"'
                f"({self.cell_coords(row_index, col_index)})"
            )
            self._process_col_helper(
                opt_name, cell_content, row_index, col_index, state, False
            )
        else:
            col_desc = self.cell_content_desc(cell_content, row_index, col_index)
            state.non_header_vals.append(col_desc)

    def _process_col_helper(
        self,
        canonical_name: str,
        cell_content,
        row_index: int,
        col_index: int,
        layout_state: ColLayoutDetectionState,
        required_col: bool,
    ):

        obs_cols_dict: Dict[str, List[int]] = (
            layout_state.obs_req_cols if required_col else layout_state.obs_opt_cols
        )

        # if column name is already observed, build a helpful error message
        if canonical_name in obs_cols_dict:
            cols: List[int] = obs_cols_dict[canonical_name]
            subcategory = f'"{canonical_name}"'

            details = []
            if len(cols) == 1:
                # since this is the first duplication, log an error for the initial occurrence
                # of this column name
                details.append(self.cell_coords(row_index, cols[0]))
            details.append(self.cell_coords(row_index, col_index))
            add_errors(
                self.import_uuid,
                DuplicateColumnError(subcategory=subcategory, details=details),
            )

        # otherwise, save state re: where we found it
        else:
            layout_state.layout[canonical_name] = col_index
            if canonical_name != cell_content:
                self.obs_col_names[canonical_name] = cell_content

        obs_cols_dict[canonical_name].append(col_index)

    def _get_raw_value(self, row, col_name):
        """
        Gets the raw value from the table, regardless of file format.

        :param row: an iterable of columns in this row of the file
        :param col_name: the canonical name of the column whose value should be read
        :return: the value, or None if the col with col_name isn't in the file
        """
        col_index = self.column_layout.get(col_name, None) if col_name else None

        if col_index is None:
            return None

        cell = row[col_index]
        return self._raw_cell_value(cell)


@dataclass(frozen=True)
class MeasurementParseRecord:
    """
        A record resulting from parsing a single value or set of related values from an import
        file.
        This object should be flexible enough to capture the full level of detail needed to
        construct a Measurement & MeasurementValue from any file, though many formats won't require
        this level of detail.

        :param loa_name: the name of the line or assay this record applies to. Whether it
        matches to a
            line or to an assay is determined subsequently during the import.
        :param mtype_name: the name of the MeasurementType this record measures.  E.g. a Uniprot
            accession ID or PubChem Compound ID
        :param format: the Measurement.Format instance that describes layout of the data in this
        record
        :param meta: a dict that maps type_name => value for any metadata associated with this
        record
        :param data: an array of data to store in the database, packed according to "format"
        :param x_unit_name: the string name for MeasurementUnits on the x-axis in this record (e.g.
            "hours" is built-in)
        :param y_unit_name: the string name for MeasurementUnits for the y-axis in this record (
        e.g.
            "g/L" is built in)
        :param src_ids: an iterable of identifiers for file locations this record was sourced from.
            This is important input for constructing helpful error messages if problems occur
            downstream (e.g. in resolving values from the file with reference databases)

        Compare with RawImportRecord from the legacy import.  Major differences are:
        1. "Kind" is removed
        2. No deep copying -- client code retains control
        3. Units have been added
        4. src_ids is added
        5. x and y are split instead of being merged early during the import for later unmerge
        """

    loa_name: str
    mtype_name: str
    format: str
    data: List[List[numbers.Number]]
    x_unit_name: str
    y_unit_name: str
    # data source(s) within the file for this measurement...for tabular data, an iterable of
    # int row nums or string ranges, e.g. ('1-3', 5, 24). Used to construct helpful / precise
    # error messages
    src_ids: [Tuple[str, ...]]

    def __str__(self):
        return (
            f"MeasurementParseRecord(loa_name={self.loa_name}, "
            f"mtype_name= {self.mtype_name}, format={self.format}, data={self.data}, "
            f"({self.x_unit_name}, {self.y_unit_name})"
        )

    def time(self):
        return self.data[0][0]

    def to_json(self):
        return {
            "loa_name": self.loa_name,
            "measurement_name": self.mtype_name,
            "y_unit_name": self.y_unit_name,
            "x_unit_name": self.x_unit_name,
            "format": self.format,
            "data": self.data,
            "src_ids": self.src_ids,
        }


@dataclass(init=False, frozen=True)
class FileParseResult:
    """
        Standard format for capturing the results of parsing a file for import.  Note this doesn't
        imply that the import can necessarily succeed, just that the baseline of required data in
        the file has been read and verified for the expected basic data type (e.g. numeric,
        string).
    """

    # True if the file contained at least one time value
    any_time: bool = field(compare=False)

    # True if a time value was found corresponding to each MeasurementParseRecord read from file
    # Since parsed values may be packed differently for for different use cases, it's simpler
    # for the parser to make this determination.
    has_all_times: bool = field(compare=False)

    # True if every record parsed from file had associated units
    has_all_units: bool = field(compare=False)

    series_data: Sequence[MeasurementParseRecord]

    #  a human-readable identifier for the file portion(s) identified by each
    #  MeasurementParseRecord's src_ids. For example, for tabular data, "row" is
    #  often used
    record_src: str

    # set of unique line or assay names found in the file
    line_or_assay_names: Set[str] = field(compare=False)

    # set of unique mtype identifiers (strings) found in the file
    mtypes: Set[str] = field(compare=False)

    # set of unique unit names (strings) found in the file or implicit in the format
    units: Set[str] = field(compare=False)

    def __init__(
        self,
        series_data: Iterable[MeasurementParseRecord],
        record_src: str,
        any_time: bool,
        has_all_times: bool,
    ):
        object.__setattr__(self, "series_data", series_data)
        object.__setattr__(self, "record_src", record_src)
        object.__setattr__(self, "any_time", any_time)
        object.__setattr__(self, "has_all_times", has_all_times)

        # compute unique line / assay names, units, measurement types from the parse records
        line_or_assay_names: Set[str] = set()
        units: Set[str] = set()
        mtypes: Set[str] = set()

        has_all_units = True

        for record in self.series_data:
            line_or_assay_names.add(record.loa_name)
            units.add(record.x_unit_name)
            units.add(record.y_unit_name)
            mtypes.add(record.mtype_name)

            if not record.x_unit_name or not record.y_unit_name:
                has_all_units = False

        object.__setattr__(self, "line_or_assay_names", frozenset(line_or_assay_names))
        object.__setattr__(self, "has_all_units", has_all_units)
        object.__setattr__(self, "mtypes", frozenset(mtypes))
        object.__setattr__(self, "units", frozenset(units))


class ExcelParserMixin:
    def parse(self, file):
        """
        Parses the input file as an Excel workbook.
        :param file: a file-like object
        :return: if parsing was successful, the FileParseResult read from file, otherwise None
        :raise: OSError If the file can't be opened or EDDImportError if the file format or content
        is bad
        """
        wb = load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        logger.debug("In parse(). workbook has %d sheets" % len(wb.worksheets))
        if not wb.worksheets:
            raise_errors(self.import_uuid, EmptyFileError())
        elif len(wb.worksheets) > 1:
            sheet_name = wb.sheetnames[0]
            count = len(wb.worksheets) - 1
            intro = _(
                'Only the first sheet in your workbook, "{sheet}", was processed. '
            ).format(sheet=sheet_name)
            if count > 2:
                postfix = _("All other sheets were ignored ({count}).").format(
                    count=count
                )
            else:
                postfix = _('The other sheet "{name}" was ignored.').format(
                    count=count, name=wb.sheetnames[1]
                )
            msg = intro + postfix
            warnings(self.import_uuid, IgnoredWorksheetWarning(details=[msg]))
        worksheet = wb.worksheets[0]
        return self._parse_rows(worksheet.iter_rows())

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file

        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped if the content
            was a string
        """
        val = cell.value
        if isinstance(val, string_types):
            return val.strip()
        return val


class CsvParserMixin:
    def parse(self, file):
        """
        Parses the input file a CSV file

        :param file: a file-like object
        :return: if parsing was successful, the FileParseResult read from file,
            otherwise None
        :raise: OSError If the file can't be opened or EDDImportError if the file format
            or content is bad
        """
        reader = csv.reader(file)
        return self._parse_rows(reader)

    def _raw_cell_value(self, cell):
        val = cell
        if isinstance(val, string_types):
            return val.strip()
        return val


def build_src_summary(sources, convert_ints=False):
    """
    Takes a list of data source descriptions, e.g. integer rows in a file, and condenses it by
    merging consecutive integers into a string representing a range.

    For example, [1, 2, 3, 5] => ["1-3", 5]. Non-integers are simply copied to the result,
    so for example [1, 2, 3, "18-36", 42] =>  ["1-3", "18-36", 42].

    :param convert_ints: True to convert any single integers in the result to strings, False
        to leave them as ints
    :return: a list of source ranges, where ranges are strings and any single sources are
        either ints or strings as determined by convert_ints
    """
    src_ranges = []

    range_start = 0
    for i in range(1, len(sources)):
        curr = sources[i]
        prev = sources[i - 1]

        if not isinstance(curr, int) or (not isinstance(prev, int)) or curr != prev + 1:
            if i == range_start + 1:
                val = (
                    sources[range_start]
                    if not convert_ints
                    else str(sources[range_start])
                )
                src_ranges.append(val)
            else:
                src_ranges.append(f"{sources[range_start]}-{prev}")
            range_start = i

    last = sources[len(sources) - 1]
    if range_start == len(sources) - 1:
        src_ranges.append(last if not convert_ints else str(last))
    else:
        src_ranges.append(f"{sources[range_start]}-{last}")
    return src_ranges
