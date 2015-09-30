#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Read in data from a spreadsheet.

import logging
import openpyxl as px

from six import string_types


logger = logging.getLogger(__name__)


# Datablock dimesions for 96 well plate
DATA_BLOCK_SIZE_ROW = 12
DATA_BLOCK_SIZE_COL = 8


def process_spreadsheet(input_file_path):
    """Processes a spreadsheet at a given path, returning the datablocks found.

    Datablocks come in the form (read,block_type,data_block)

    read is a string identifying which data set the block is associated with
    block_type is a string identifying the type of metadata block
        Data - A 96 well plate grid of measurements.
        ControlData - A 96 well plate grid of measurements, for a control read.
        Metadata - A list of key-value pairs, applying globally to the read.
        other - A 96 well plate grid of some well-specific metadata item - label.
    datablock - A 1D List or Dictionary

    """
    W = px.load_workbook(input_file_path, use_iterators=True, read_only=True)

    data_blocks = []
    for sheet in W:
        data_block_indices = _locate_data_blocks(sheet)
        for index in data_block_indices:
            data_blocks.append(_collect_data_block(sheet, index))

    return data_blocks


# *** *** *** Internals *** *** ***


def _locate_data_blocks(sheet):
    """Scans a sheet for a data_block tag (96_wellplate_read_*) and returns all such indices"""
    data_block_indices = []
    data_block_indicator_prefix = u"96_wellplate_read_"

    for row in sheet.rows:
        for cell in row:
            if (cell.value and isinstance(cell.value, string_types) and
                    cell.value.startswith(data_block_indicator_prefix)):
                if (cell.value == 'Study_name' or cell.value == 'Study_description'):
                    continue
                data_block_indices.append((cell.row, cell.column))
    logger.debug('data_block_indices: %s' % data_block_indices)
    return data_block_indices


# Convert from letter(s) to a number
# ORIG did not deal with multiletter #column = ord(data_block_index[1])-ord('A')+1
def _convert_column_letters_to_integer(letters):
    """converts excel column letters to index number"""
    number = 0

    # TODO: REFACTOR: reduce
    if len(letters) > 1:
        final_letter = letters[len(letters)-1]
        number += 26 * (len(letters)-1)
        number += ord(final_letter)-ord('A')+1
    else:
        number = ord(letters)-ord('A')+1
    return number


def _collect_data_block(sheet, data_block_index):
    """Collects the data associated with each datablock index"""
    data = []
    row = data_block_index[0]
    column = _convert_column_letters_to_integer(data_block_index[1])
    x = y = 0

    wellplate_read = sheet.cell(row=row, column=column).value
    label = sheet.cell(row=row+1, column=column).value
    logger.debug('wellplate_read: %s' % wellplate_read)
    logger.debug('label: %s' % label)

    if label == u'Metadata':
        x += 2
        while True:
            if sheet.cell(row=row+x, column=column+y).value == 'MetadataEnd':
                break
            data.append((sheet.cell(row=row+x, column=column+y).value,
                         sheet.cell(row=row+x, column=column+y+1).value))
            x += 1
        data_block = dict(data)
        logger.debug('%s' % data_block)
    else:
        row += 1
        column += 1
        logger.debug(sheet.cell(row=row+x, column=column+y).value)
        for y in range(DATA_BLOCK_SIZE_ROW):
            for x in range(DATA_BLOCK_SIZE_COL):
                data.append(sheet.cell(row=row+x, column=column+y).value)
        data_block = data
    return (wellplate_read, label, data_block)

# TODO: place former __main__ code in a test
