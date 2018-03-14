# coding: utf-8
"""
Convenience methods wrapping openpyxl module for handling .xlsx file I/O.
(Despite the module name, this doesn't actually do any parsing of the file
format, but it attempts to intelligently interpret the content.)
"""

import sys

from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
from openpyxl import Workbook
from six import string_types


def export_to_xlsx(table, headers=None, title="Exported table", file_name=None):
    """
    Create a simple Excel workbook from the given table and optional headers.
    """
    assert_is_two_dimensional_list(table)
    # XXX https://bitbucket.org/openpyxl/openpyxl/issue/375/use-save_virtual_workbook-and-optimized
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = title
    if (headers is not None):
        assert (len(table) == 0) or (len(headers) == len(table[0]))
        ws.append(headers)
    for row in table:
        ws.append(row)
    if (file_name is None):
        return save_virtual_workbook(wb)
    else:
        wb.save(file_name)
        return file_name


def _get_sheets_from_workbook(wb, name, worksheet_id):
    ws = None
    if name is not None:
        ws = [sheet for sheet in wb.worksheets if sheet.title == name]
        if len(ws) == 0:
            raise KeyError(f"Cannot find worksheet named {name}")
    elif worksheet_id is not None:
        ws = [wb.worksheets[worksheet_id]]
    else:
        ws = wb.worksheets
    return ws


def import_xlsx_tables(
        file,
        column_labels=None,
        column_search_text=None,
        worksheet_id=None,
        worksheet_name=None,
        enforce_non_blank_cells=False,
        expect_numeric_data=False):
    """
    Process an excel file and extract the coherent table(s).  This will attempt
    to use several heuristics to identify what parts of the worksheet(s) contain
    a table.  The most reliable method is to specify column labels to search for,
    at least two of which must be found.  A more approximate method is to search
    for a specific keyword in the headers.  If neither of these are given, any
    table-like block of cells will be extracted.  In all cases, a row where the
    first expected cell is blank signals the end of a contiguous table, but
    additional rules apply for the fully-automatic extraction.

    :param file: Excel 2007+ (.xlsx) file name or handle
    :param column_labels: specific columns names to extract (not case sensitive)
    :param column_search_text: string to search for as column header (not case
        sensitive)
    :param worksheet_id: index of worksheet to extract from
    :param worksheet_name: name of worksheet to extract from
    :param enforce_non_blank_cells: treat rows containing blank internal cells
        as the end of a possible table
    :param expect_numeric_data: indicates that tables should contain numeric
        values (as opposed to text)
    :returns: dict of results with single 'worksheets' key (suitable for JSON
        export)
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    if (len(wb.worksheets) == 0):
        raise ValueError("No worksheets found in Excel file!")
    ws = _get_sheets_from_workbook(wb, worksheet_name, worksheet_id)
    ws_results = []
    for i_sheet, ws_ in enumerate(ws):
        try:
            ws_results.append(
                _find_table_of_values(
                    worksheet=ws_,
                    column_labels=column_labels,
                    column_search_text=column_search_text,
                    enforce_non_blank_cells=enforce_non_blank_cells,
                    expect_numeric_data=expect_numeric_data))
        except ValueError as e:
            if len(ws) > 1:
                continue
            else:
                raise
    # FIXME we can do better than this...
    if (len(ws_results) == 0):
        raise ValueError(
            "No valid table-like blocks of contiguous cells found "
            "in the file.  Please make sure your spreadsheet follows format "
            "restrictions."
        )
    return {"worksheets": ws_results}


def import_xlsx_table(*args, **kwds):
    """
    Identical to import_xlsx_tables, but only returns a single table dict (or
    raises an error if more than one was found).
    """
    result = import_xlsx_tables(*args, **kwds)
    tables = result['worksheets'][0]
    if (len(result['worksheets']) > 1) or (len(tables) > 1):
        raise ValueError("Multiple tables found.")
    return tables[0]


# XXX This might be generalizable to other applications (e.g. processing
# various text files), since it can accept a Python list-of-lists as input
def _find_table_of_values(
        worksheet,
        column_labels=None,
        column_search_text=None,
        expect_numeric_data=False,
        minimum_number_of_data_rows=2,
        maximum_number_of_tables=sys.maxsize,
        enforce_non_blank_cells=False):
    """
    Scan a worksheet for a block of cells resembling a regular table structure,
    using optional clues about content.  Returns a list of dicts with separate
    keys/value pairs for headers and actual data.  (The header list may be empty
    if no key was defined and the first row contains numeric data.)
    """
    rows = worksheet
    if (not isinstance(worksheet, list)):
        rows = worksheet_as_list_of_lists(worksheet)
    else:
        assert_is_two_dimensional_list(rows)
    n_rows = len(rows)
    # We assume these are case insensitive, so Joe Scientist doesn't need to
    # learn about Caps Lock.
    if (column_search_text is not None):
        column_search_text = column_search_text.lower()
    if (column_labels is not None):
        column_labels = {cl.lower() for cl in column_labels}
    possible_tables = []

    row_index = 0
    while (row_index < n_rows):
        row = rows[row_index]
        row_index += 1
        n_values = number_of_non_blank_cells(row)
        if n_values >= 2:
            if column_labels is not None:
                column_indices = []
                headers = []
                for i_cell, value in enumerate(row):
                    if not isinstance(value, string_types):
                        continue
                    if value.lower() in column_labels:
                        column_indices.append(i_cell)
                        headers.append(value)
                if len(headers) >= 2:
                    table = []
                    while row_index < n_rows:
                        row = [rows[row_index][k] for k in column_indices]
                        row_index += 1
                        # stop when we hit a row where (at least) the first cell is blank
                        if (row[0] is None):
                            break
                        table.append(row)
                    if (len(table) == 0):
                        raise ValueError(
                            "The column labels '%s' were found in the "
                            "worksheet, but no recognizable table of values was associated "
                            "with them." % ";".join(headers)
                        )
                    return [{
                        "headers": headers,
                        "values": table,
                    }]
            elif column_search_text is not None:
                headers = []
                table = []
                # looking for a specific column header in the row
                for i_cell, value in enumerate(row):
                    if value is None:
                        continue
                    if isinstance(value, string_types) and column_search_text in value.lower():
                        headers = row[i_cell:]
                        while (row_index < n_rows):
                            row = rows[row_index][i_cell:]
                            row_index += 1
                            if (row[0] is not None):
                                table.append(list(row))
                            else:
                                break
                        break
                if len(headers) > 0:
                    if len(table) == 0:
                        raise ValueError(
                            "The search text '%s' was found in the "
                            "worksheet, but no recognizable table of values was associated "
                            "with it." % column_search_text
                        )
                    return [{
                        "headers": headers,
                        "values": table,
                    }]
            else:
                contiguous_rows = []
                contiguous_rows.append(row)
                row_index_inner = row_index
                # scan ahead in the table for additional rows that are non-blank, and collect them
                while (row_index_inner < n_rows):
                    nb_row = rows[row_index_inner]
                    row_index_inner += 1
                    nb_values = number_of_non_blank_cells(nb_row)
                    if nb_values > 0:
                        contiguous_rows.append(nb_row)
                    else:
                        break

                # If we only found one row, it might be a headerless run of values.
                # But if it doesn't contain anything that looks numeric, forget it.
                if len(contiguous_rows) < 2:
                    n_numeric = number_of_numerical_cells(contiguous_rows[0])
                    if n_numeric < 1:
                        row_index = row_index_inner
                        continue  # Continue outer while loop - go to next chunk

                first_non_empty_column = None
                last_non_empty_column = None
                # If some rows begin or end before or after others, we'll consider the table
                # to have 'irregular row sizes'.
                irregular_row_sizes = False
                found_blank_cells = False
                for c_rows_index, c_row in enumerate(contiguous_rows):

                    first_non_empty_cell = None
                    last_non_empty_cell = None
                    for c_cell_index, c_cell in enumerate(c_row):
                        if c_cell is None:
                            # If we've already found one non-empty cell,
                            # and the row is still continuing, count it as a
                            # blank cell inside the table.
                            if first_non_empty_cell is not None:
                                found_blank_cells = True
                            continue
                        if first_non_empty_cell is None:
                            first_non_empty_cell = c_cell_index
                        else:
                            first_non_empty_cell = min(first_non_empty_cell, c_cell_index)
                        if last_non_empty_cell is None:
                            last_non_empty_cell = c_cell_index
                        else:
                            last_non_empty_cell = max(last_non_empty_cell, c_cell_index)

                    if first_non_empty_column is not None:
                        if first_non_empty_column != first_non_empty_cell:
                            first_non_empty_column = min(first_non_empty_column,
                                                         first_non_empty_cell)
                            irregular_row_sizes = True
                    else:
                        first_non_empty_column = first_non_empty_cell

                    if last_non_empty_column is not None:
                        if last_non_empty_column != last_non_empty_cell:
                            last_non_empty_column = max(last_non_empty_column, last_non_empty_cell)
                            irregular_row_sizes = True
                    else:
                        last_non_empty_column = last_non_empty_cell

                # It would be extremely odd if we got this.
                if (first_non_empty_column is None) or (last_non_empty_column is None):
                    row_index = row_index_inner
                    continue  # Outer while loop

                # Enforcing non-blank-ness means we want a rectangular table, with no holes.
                if enforce_non_blank_cells and (irregular_row_sizes or found_blank_cells):
                    row_index = row_index_inner
                    continue  # Outer while loop

                largest_row_size = (last_non_empty_column - first_non_empty_column) + 1

                # We are not going to bother with a 'table' that is 1x1, 1x2, 2x1, 1x3, or 3x1.
                if largest_row_size * len(contiguous_rows) < 4:
                    row_index = row_index_inner
                    continue  # Outer while loop

                # We are going to push these rows in 'unfiltered', starting
                # from the first non-empty column, under the assumption that
                # empty leading cells are structurally relevant -
                # e.g. they indicate null values, or act as spacing for headers.
                tmp = []
                for c_rows_index, c_row in enumerate(contiguous_rows):
                    c_row_part = c_row[first_non_empty_column:]
                    tmp.append(list(c_row_part))

                # check that we have a reasonable number of rows in current table
                if len(tmp) > minimum_number_of_data_rows:
                    possible_tables.append(tmp)
                    row_index = row_index_inner

    if column_search_text is not None:
        raise ValueError(
            "The specified search text '%s' could not be associated "
            "with a column label in this spreadsheet.  Make sure the table you "
            "wish to extract obeys the required formatting rules." % column_search_text
        )
    if len(possible_tables) > maximum_number_of_tables:
        raise ValueError(
            "Multiple table-like blocks of cells identified in "
            "this worksheet - please specify a (partial) column label to search "
            "for, or provide a simpler file containing only the table of interest."
        )
    elif len(possible_tables) == 0:
        raise ValueError(
            "No table-like blocks of cells could be automatically "
            "identified in this worksheet."
        )
    else:
        return [
            {"headers": tmp[0], "values": tmp[1:]}
            for tmp in possible_tables
        ]


def worksheet_as_list_of_lists(ws):
    """Convert a Worksheet object to a 2D list."""
    table = []
    for row in ws.rows:
        table.append([c.value for c in row])
    return table


def assert_is_two_dimensional_list(table, allow_tuple_values=False):
    """
    Verify that what we expect to be a 2D list (i.e. a list of lists) really is
    that - with the optional exception of tuple cell values.
    """
    assert isinstance(table, list)
    for row in table:
        assert (isinstance(row, list) or isinstance(row, tuple)), row
        for cell in row:
            valid_cell = isinstance(cell, string_types) or not hasattr(cell, '__iter__')
            if (allow_tuple_values):
                assert isinstance(cell, tuple) or valid_cell, cell
            else:
                assert valid_cell, cell


def number_of_numerical_cells(row):
    """Count cells with int or float types."""
    numtypes = [isinstance(c, int) or isinstance(c, float) for c in row]
    return numtypes.count(True)


def number_of_non_blank_cells(row):
    """Count cells that are not None."""
    return [(c is not None) for c in row].count(True)
