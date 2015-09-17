#!/usr/bin/env python
# -*- coding: utf-8 -*-

## Read in data from a spreadsheet.

import openpyxl as px
import numpy as np


# Datablock dimesions for 96 well plate
DATA_BLOCK_SIZE_ROW = 12
DATA_BLOCK_SIZE_COL = 8

def process_spreadsheet(input_file_path):
	"""Processes a spreadsheet at a given path, returning the datablocks found.

	Datablocks come in the form (study,label,data_block)

	study is a string identifying which study the block is associated with
	label is a string identifying the type of metadata block
		Data - A 96 well plate grid of measurements.
		Metadata - A list of key-value pairs, applying globally to the study.
		other - A 96 well plate grid of some well-specific metadata item - label.
	datablock - A 1D List or Dictionary

	"""
	W = px.load_workbook(input_file_path, use_iterators = True, read_only=True)

	# p = W.get_sheet_by_name(name = 'Sheet1')

	data_blocks = []
	for sheet in W:
		data_block_indices = _locate_data_blocks(sheet)
		for index in data_block_indices:
			data_blocks.append( _collect_data_block(sheet, index))

	return data_blocks





""" *** *** *** Internals *** *** *** """


def _locate_data_blocks(sheet):
	"""Scans a sheet for a data_block tag (Study_*) and returns all such indices"""
	data_block_indices = []
	#data_block_indicator = u"Temperature(°C)"
	#data_block_indicator = u"Table_1"
	data_block_indicator_prefix = u"Study_"

	for row in sheet.rows:
		for cell in row:
			#if cell.value == data_block_indicator:
			if cell.value:
				if type(cell.value) == type(u' '):
					if cell.value.startswith(data_block_indicator_prefix):
						if (cell.value == 'Study_name' or cell.value == 'Study_description'):
							continue
						data_block_indices.append( (cell.row,cell.column) )
						# print(data_block_indices)
						# exit(1)

	print('data_block_indices: '+ str(data_block_indices))
	# exit(1)

	return data_block_indices

# Convert from letter(s) to a number
# ORIG did not deal with multiletter #column = ord(data_block_index[1])-ord('A')+1 
def _convert_column_letters_to_integer(letters):
	"""converts excel column letters to index number"""
	number = 0

	# TODO: REFACTOR: reduce
	if(len(letters)>1):
		# print('letters: ' + str(letters))
		final_letter = letters[len(letters)-1]
		number += 26 * (len(letters)-1)
		number += ord(final_letter)-ord('A')+1
		# print number
	else:
		number = ord(letters)-ord('A')+1 
	return number

def _collect_data_block(sheet, data_block_index):
	"""Collects the data associated with each datablock index"""
	data = []
	row = data_block_index[0]
	column = _convert_column_letters_to_integer(data_block_index[1])
	x = y = 0

	study = sheet.cell( row=row,   column=column ).value
	label = sheet.cell( row=row+1, column=column ).value
	print('study: ' + str(study))
	print('label: ' + str(label))

	if(label==u'Metadata'):
		# print 'metadata...'
		x += 2
		while type( sheet.cell( row=row+x, column=column+y ) )!= type(None):
			# data.append( sheet.cell( row=row+x, column=column+y   ).value )
			# data.append( sheet.cell( row=row+x, column=column+y+1 ).value )
			data.append( (sheet.cell( row=row+x, column=column+y   ).value, \
				          sheet.cell( row=row+x, column=column+y+1 ).value ))
			# data.append(((row+x, column+y ),(row+x, column+y+1)))
			x+=1
		# print data
		# matrix = np.resize(data, [x-2,2])
		data_block = dict(data)
		print data_block
		# exit(1)
	else:
		# print 'data...'
		row+=1
		column+=1
		print (sheet.cell( row=row+x, column=column+y ).value)
		for y in range(DATA_BLOCK_SIZE_ROW):
			for x in range(DATA_BLOCK_SIZE_COL):
				# data.append( sheet.cell( row=row+x, column=column+y ).internal_value)
				data.append( sheet.cell( row=row+x, column=column+y ).value )
		data_block = data
		#data_block = np.resize(data, [DATA_BLOCK_SIZE_COL,DATA_BLOCK_SIZE_ROW])
	# import IPython
	# IPython.embed()
	# exit()
	return (study,label,data_block)

if __name__ == "__main__":
	# input_file_path = "input/template_with_one_plate.xlsx"
	# input_file_path = "input/template_with_eight_plate.xlsx"
	input_file_path = "input/example_with_one_plate.xlsx"
	process_spreadsheet( input_file_path )
