# coding: utf-8

import collections
import csv
import decimal
import logging
import numbers
import re
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from six import string_types

from ..codes import FileParseCodes
from ..utilities import ErrorAggregator, ParseError

logger = logging.getLogger(__name__)


def _has_value(*args):
    """
    Tests whether any of the provided arguments should be treated as having a valid value
    """

    def _valid_value(arg):
        if isinstance(arg, string_types):
            return bool(arg)  # assume it's already .strip()ed
        else:
            return arg is not None

    return any(filter(_valid_value, args))


class TableParser(object):
    """
    A parser for the tabular import files that allows supports semi-tolerant parsing of
    user-provided spreadsheets. Basic conditions for parsing are:

    1) Required columns can be provided in any order
    2) Optional columns will be identified if present
    3) Column headers are case-insensitive
    4) Any column whose header doesn't match those specified by the file format will be
        ignored (with a warning)
    5) Whitespace parsing is tolerant, e.g. different number and type of whitespace is allowed in
        column headers than what's suggested, though space is still needed in between words if
        column headers are defined with internal whitespace.
    """
    # TODO: capture common method parameters as a class to simplify interface
    def __init__(self, req_cols, opt_cols=None, value_opt_cols=None, numeric_cols=None,
                 supported_units=None, aggregator=None):
        """
        Initializes the parser using client-provided lists of required and optional column headers.

        :param req_cols: an iterable of human-readable strings containing required column
            headers. The parser will take care of tolerating variations in optional leading,
            trailing, and internal whitespace.
        :param opt_cols: an iterable of human-readable strings with optional column headers.
            Subject to the same processing as req_cols.
        :param value_opt_cols: an optional iterable of column headers from either or both of
            req_cols & opt_cols. Any col header in value_opt_cols will have missing
            values tolerated in any row.  Any column headers not listed here are by default
            required to have a value in every row, assuming the column is present.
        :param numeric_cols: an  optional list of column names for columns that will be verified
            during parsing to have numeric content.
        :param aggregator: an optional main.importer2.ErrorAggregator instance to report errors
            and warnings to
        """
        super(TableParser, self).__init__()

        # TODO: add support for vector input
        # TODO: add optional column aliases?
        self.aggregator = aggregator if aggregator else ErrorAggregator()

        # keep inputs to use as keys and as human-readable for use in err messages
        self.req_cols = req_cols
        self.opt_cols = opt_cols if opt_cols is not None else set()
        self.value_opt_cols = value_opt_cols if value_opt_cols is not None else set()
        self.numeric_cols = numeric_cols if numeric_cols is not None else set()
        self.supported_units = supported_units if supported_units is not None else {}

        # maps canonical col name -> observed text (e.g. maybe including whitespace, mixed case,
        # alias)
        self.obs_col_names = {}
        self.obs_col_units = {}

        self._is_excel = None

        # build up regex patterns for each supported column header that allow us to do
        # semi-tolerant matching against user-provided values.
        # maps human-readable col name => pattern
        self.req_col_patterns = self._build_col_header_patterns(req_cols)

        if opt_cols:
            self.opt_col_patterns = self._build_col_header_patterns(self.opt_cols)
        else:
            self.opt_col_patterns = []

        self._build_unit_patterns()

    @property
    def has_all_times(self):
        """
        Tests whether the parsed file contained time values for all measurements.  Incomplete time
        values should be treated as a parse error.
        """
        return False  # children should override

    def has_all_units(self):
        """
        Tests whether the parsed file contained units for all measurements
        Partial units should be treated as a parse error.
        """
        return False  # children should override

    def _build_col_header_patterns(self, col_headers):
        """
        Builds patterns to facilitate reasonably tolerant parsing of fixed string values expected
        to match during parsing.  Patterns are constructed to match case-insensitive input and
        are tolerant of leading and trailing whitespace insertions, or of added internal
        whitespace.  Internal whitespace in the input is required, but may use different
        whitespace characters or have additional whitespace added.
        """
        reg = r'^\s*{title}\s*$'
        return [
            re.compile(reg.format(title=TableParser._process_label(col_header)), re.IGNORECASE)
            for col_header in col_headers
        ]

    def _build_unit_patterns(self):
        self._unit_patterns = {}
        for col, units in self.supported_units.items():
            logger.debug(f'Building unit patterns for column "{col}": {units}')
            # note : maintaining case-sensitivity is important! SI units use case!
            vals = '|'.join([TableParser._process_label(unit) for unit in units])
            pat = re.compile(r'^\s*({vals})\s*$'.format(vals=vals))
            self._unit_patterns[col] = pat

    @staticmethod
    def _process_label(s):
        s = re.sub(r'\s+', ' ', s)  # collapse multiple whitespace chars into a single space char
        s = re.escape(s)  # escape any regex chars in the input (also escapes space char)
        return re.sub(r'\\ ', r'\s+', s)  # replace escaped space with space regex

    def parse_excel(self, file):
        """
        Parses the input file as an Excel workbook.
        :param file: a path-like object identifying the file location
        :return: if parsing was successful, the number of valid rows read from file, otherwise None
        :raise: OSError If the file can't be opened or ParseError if the file format or content
        is bad
        """
        self._is_excel = True
        wb = load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        logger.debug('In parse(). workbook has %d sheets' % len(wb.worksheets))
        if not wb.worksheets:
            self.aggregator.add_error(FileParseCodes.EMPTY_FILE)
            raise ParseError(self.aggregator)
        elif len(wb.worksheets) > 1:
            sheet_name = wb.sheetnames[0]
            msg = (f'Only the first sheet in your workbook, "{sheet_name}", was processed.  '
                   f'All other sheets will be ignored.')
            self.aggregator.add_warning(FileParseCodes.IGNORED_WORKSHEET,
                                        occurrence=msg)
        worksheet = wb.worksheets[0]
        return self._parse(worksheet.iter_rows())

    def parse_csv(self, file):
        """
        Parses the input file as CSV.
        :param file: a path-like object identifying the file location
        :return: if parsing was successful, the number of valid rows read from file, otherwise None
        :raise: OSError If the file can't be opened, or ParseError if the file format or content is
        bad
        """
        self._is_excel = False
        reader = csv.reader(file)
        return self._parse(reader)

    def _parse(self, rows_iter):
        """
        A workhorse method that performs parsing, independent of how the file content is stored
        (CSV or Excel).
        :return: the number of non-whitespace rows parsed
        :raise ParseError: if the file format or content was bad
        """
        # Clear out state from any previous use of this parser instance
        self.column_layout = None
        obs_required_cols = set()

        # loop over rows
        for row_index, cols_list in enumerate(rows_iter):
            logger.debug(f'Parsing row {row_index+1}')

            # identify columns of interest first by looking for required labels
            if not self.column_layout:
                self.column_layout = self._parse_col_layout(cols_list, row_index,
                                                            obs_required_cols)

                if self.column_layout:
                    self._verify_layout(self.aggregator, row_index)

            # if column labels have been identified, look for data
            else:
                self._parse_row(cols_list, row_index)

        # if we got all the way through the file without finding a single required column header,
        # it's badly formatted
        if not self.column_layout:
            logger.debug(f'Required columns were not read {self.req_cols}')
            self.aggregator.add_error(FileParseCodes.MISSING_REQ_COL_HEADER,
                                      ', '.join(self.req_cols))

        if self.aggregator.errors:
            raise ParseError(self.aggregator)

        return self.measurement_count

    def cell_content_desc(self, content, row_index, col_index):
        """
        Builds a succinct, human-readable description of the cell content and location
        :param content: the cell content
        :param row_index: index into the file where this row is located
        :param col_index: index into the file where this column is located
        :return: a string description of the cell content and location
        """
        return f'"{content}" ({self.cell_coords(row_index, col_index)})'

    def _parse_col_layout(self, row, row_index, obs_required_cols):
        """
        Scans a row of the file to see if it contains required column headers that define the file
        layout.
        :param row: the row to inspect for column headers
        :param row_index: the index into the file of the row being checked for required col headers
        :param obs_required_cols: a set to populate with required col names observed so far in the
            file.  Since detection of any required column header counts as a match, this parameter
            is essentially just an optimization to prevent creating a new set each time we try
            to read a row in poorly-formatted files.
        :return: the column layout if required columns were found, or None otherwise
        """
        layout = {}  # maps col name to col index
        importer = self.aggregator
        req_cols_set = set(self.req_cols)

        ###########################################################################################
        # loop over columns in the current row
        ###########################################################################################
        obs_opt_cols = set()
        non_header_vals = []
        for col_index, cell in enumerate(row):
            cell_content = self._raw_cell_value(cell)

            # skip this cell if it has no non-whitespace content
            # (strings are stripped by _get_raw_value())
            if not cell_content:
                continue

            # ignore non-string cells since they can't be the column headers we're looking for
            if not isinstance(cell_content, string_types):
                if cell_content is not None:
                    col_desc = self.cell_content_desc(cell_content, row_index, col_index)
                    non_header_vals.append(col_desc)
                continue

            #######################################################################################
            # check whether column label matches one of the canonical column names specified by
            # the format
            #######################################################################################
            self._process_col_name(cell_content, row_index, col_index, non_header_vals,
                                   obs_required_cols, obs_opt_cols, layout)

        # if at least the required columns were found, consider this a successful read
        if obs_required_cols == req_cols_set:
            for bad_val in non_header_vals:
                importer.add_warning(FileParseCodes.COLUMN_IGNORED, occurrence=bad_val)
            logger.debug(f'Found all {len(self.req_cols)} required column headers in row '
                         f'{row_index+1}')
            return layout

        # if some--but not all--required columns were found in this row, consider this a
        # poorly-formatted file
        if obs_required_cols:
            missing_cols = req_cols_set.difference(obs_required_cols)
            logger.debug(f'Required column headers missing: {missing_cols}')
            importer.raise_errors(FileParseCodes.MISSING_REQ_COL_HEADER, occurrences=missing_cols)
        else:
            importer.add_warnings(FileParseCodes.IGNORED_VALUE_BEFORE_HEADERS, non_header_vals)

        return None

    def _process_col_name(self, cell_content, row_index, col_index, non_header_vals,
                          obs_req_cols, obs_opt_cols, layout):
        """
        Process the string content of a cell when looking for columns
        :param cell_content: the string content (non empty, non-whitespace)
        :param non_header_vals: a list of non column header values read from the file before
        any valid column header is found
        :param obs_req_cols: a set of required column headers already observed in the row
        :param obs_opt_cols: a set of optional columns headers already observed in the row
        :param layout: a dict that maps col name to the column index where it was observed
        """
        req_name = self._parse_col_header(self.req_col_patterns, self.req_cols, row_index,
                                          col_index, cell_content)
        opt_name = self._parse_col_header(self.opt_col_patterns, self.opt_cols, row_index,
                                          col_index, cell_content)
        agg = self.aggregator
        if req_name:
            logger.debug(f'Found required column "{req_name}"'
                         f'({self.cell_coords(row_index, col_index)})')

            if req_name in obs_req_cols:
                col_desc = self.cell_content_desc(cell_content, row_index, col_index)
                agg.add_error(FileParseCodes.DUPLICATE_COL_HEADER, occurrence=col_desc)
            else:
                obs_req_cols.add(req_name)
                layout[req_name] = col_index
                if req_name != cell_content:
                    self.obs_col_names[req_name] = cell_content
        elif opt_name:
            logger.info(f'Found optional column "{opt_name}".'
                        f'({self.cell_coords(row_index, col_index)})')
            if opt_name in obs_opt_cols:
                col_desc = self.cell_content_desc(cell_content, row_index, col_index)
                agg.add_error(FileParseCodes.DUPLICATE_COL_HEADER, occurrence=col_desc)
            else:
                obs_opt_cols.add(opt_name)
                layout[opt_name] = col_index
                if cell_content != opt_name:
                    self.obs_col_names[opt_name] = cell_content
        else:
            col_desc = self.cell_content_desc(cell_content, row_index, col_index)
            non_header_vals.append(col_desc)

    def _raw_cell_value(self, cell):
        """
        Gets the raw cell value in whatever format it was stored in the file
        :param cell: the cell
        :return: the cell value, with leading and trailing whitespace stripped if the content
        was a string
        """
        if self._is_excel:
            val = cell.value
        else:
            val = cell

        if isinstance(val, string_types):
            return val.strip()

        return val

    def _get_raw_value(self, row, col_name):
        """
        Gets the raw value from the table, regardless of file format.
        :param row: an iterable of columns in this row of the file
        :param col_name: the canonical name of the column whose value should be read
        :return: the value, or None if the col with col_name isn't in the file
        """
        col_index = self.column_layout.get(col_name, None) if col_name else None

        if col_index is None:
            return None

        cell = row[col_index]
        return self._raw_cell_value(cell)

    @property
    def measurement_count(self):
        """
        Gets the number of valid measurements read from the file
        """
        # children must implement
        raise NotImplementedError()

    def obs_col_name(self, canonical_name):
        """
        Gets the observed column label from the file corresponding to the canonical name understood
        by the parser.
        :param canonical_name: the canonical column name.
        :return: the observed column name, or the canonical column name if the column wasn't found
            in the file.
        """
        return self.obs_col_names.get(canonical_name, canonical_name)

    # TODO: consider (optional) mtype ID pattern verification
    def _parse_and_verify_val(self, val, row_index, col_name):
        """
        Tests a parsed value, and logs an error message if it the observed value doesn't meet
        expectations. Verifies that:
        1) The value is present if required
        2) It's a scalar number, if required
        3) If it's a unit, that it falls within the set of supported units
        :param val: the value read from file
        :param row_index: the index of the file row the value is from
        :param col_name: the canonical name for this column, rather than the one observed in
        the file.
        :return: the value, if it was valid, or None if it failed one ore more validation criteria
        """
        importer = self.aggregator
        col_index = self.column_layout.get(col_name, None)

        # if parser is asking for a column not found in the file, end early.
        # if it's required, an error will already have been recorded
        if col_index is None:
            logger.warning(f'Column "{col_name}" not found in file')
            return None

        # if value is missing, but required, log an error
        if val is None:
            if col_name not in self.value_opt_cols:
                importer.add_error(FileParseCodes.MISSING_REQ_VALUE,
                                   subcategory=self.obs_col_name(col_name),
                                   occurrence=self.cell_coords(row_index, col_index))
            return val

        # if observed value is a string,
        if isinstance(val, string_types):
            if (not val) and col_name not in self.value_opt_cols:
                importer.add_error(FileParseCodes.MISSING_REQ_VALUE,
                                   subcategory=self.obs_col_name(col_name),
                                   occurrence=self.cell_coords(row_index, col_index))
                return None

            # parse / verify expected numeric values... when read from CSV, they'll have to be
            # parsed
            if col_name in self.numeric_cols:
                return self._parse_num(val, col_name, row_index, col_index)

            return val

        # assumption (for now) is that value is numeric. should work for both vector (
        # 1-dimensional) and numeric inputs
        if not isinstance(val, numbers.Number):
            importer.add_error(FileParseCodes.INVALID_VALUE,
                               subcategory=self.obs_col_name(col_name),
                               occurrence=self.cell_coords(row_index, col_index))

        return val

    def _parse_num(self, token, col_name, row_index, col_index):
        try:
            return decimal.Decimal(token)
        except (decimal.InvalidOperation, decimal.Clamped):
            self.aggregator.add_error(FileParseCodes.INVALID_VALUE,
                                      subcategory=self.obs_col_name(col_name),
                                      occurrence=self.cell_content_desc(token, row_index,
                                                                        col_index))
            return None

    def _parse_col_header(self, col_patterns, col_names, row_index, col_index, cell_content):
        """
        Tests the cell content against the canonical column definitions, and extracts the
        canonical column name and optional observed units if it matches.
        :param col_patterns: the iterable of column name patterns to match against
        :param col_names: an iterable of canonical column names, parallel structure to
        col_patterns
        :param row_index: the row index currently being parsed
        :param col_index: the column index currently being parsed
        :param cell_content: the content of the current cell
        """
        # loop over all column patterns, even those already matched.  Otherwise we can't detect
        # duplicates.
        for col_pattern, col_name in zip(col_patterns, col_names):
            if col_pattern.match(cell_content):
                return col_name

        return None

    def _verify_layout(self, importer, header_row_index):
        pass  # children may optionally implement

    def _parse_row(self, cols_list, row_index):
        raise NotImplementedError()  # children must implement

    @property
    def series_data(self):
        raise NotImplementedError()  # children must implement

    @property
    def units(self):
        raise NotImplementedError()  # children must implement

    @property
    def mtypes(self):
        raise NotImplementedError()  # children must implement

    @property
    def line_or_assay_names(self):
        raise NotImplementedError()

    @staticmethod
    def cell_coords(row_index, col_index):
        col_letter = get_column_letter(col_index+1)
        return f'{col_letter}{row_index+1}'

    def _verify_required_val(self, value, row_index, col_index, col_title):
        if value is None or (isinstance(value, string_types) and not value.strip()):
            self.aggregator.add_error(FileParseCodes.MISSING_REQ_VALUE, subcategory=col_title,
                                      occurrence=self.cell_coords(row_index, col_index))


# TODO: in a later version, resolve with RawImportRecord from the older import...only significant
# differences here are that "kind" is removed, no deep copying, and units & src have been added.
# Best to keep things separate for now.  TODO: update this comment prior to commit
class MeasurementParseRecord(object):
    """
    A record resulting from parsing a single Measurement from an import file.  This object should
    be flexible enough to capture the full level of detail needed to construct a Measurement from
    any file, though many formats won't require this level of detail.

    Compare with RawImportRecord from the legacy import.  Major differences are:
    1. "Kind" is removed
    2. No deep copying -- client code retains control
    3. Units have been added
    4. src_id is added
    """
    def __init__(self, **kwargs):
        self.line_or_assay_name = kwargs.get('line_or_assay_name', None)
        self.mtype_name = kwargs.get('mtype_name', None)
        self.data = kwargs.get('data', None)
        self.metadata_by_name = kwargs.get('meta', None)
        self.units_name = kwargs.get('units_name', None)

        # data source for this measurement...often a row num if input is Excel
        self.src_id = kwargs.get('src_id', None)

    def to_json(self):
        return {
            "line_or_assay_name": self.line_or_assay_name,
            "measurement_name": self.mtype_name,
            "metadata_by_name": self.metadata_by_name,
            "units_name": self.units_name,
            "data": self.data,
        }


class GenericImportParser(TableParser):
    """
    Parser for EDD's "Generic" import file, a normalized, a simple, accessible,
    and machine-readable tabular format designed for automated data import.
    """
    def __init__(self, aggregator=None):

        super(GenericImportParser, self).__init__(
            req_cols=['Line Name', 'Time', 'Measurement Type', 'Value', 'Units'],
            numeric_cols=['Time', 'Value'],
            aggregator=aggregator)
        self._measurements = []
        self.unique_units = set()
        self.unique_mtypes = set()
        self.unique_line_or_assay_names = set()

    def _verify_layout(self, importer, header_row_index):
        pass

    def _parse_row(self, cells_list, row_index):
        # extract raw values from in-use cols in this row
        # TODO: should rename col header for clarity if this could actually be assay name
        line_or_assay_name = self._get_raw_value(cells_list, 'Line Name')
        mtype = self._get_raw_value(cells_list, 'Measurement Type')
        val = self._get_raw_value(cells_list, 'Value')
        time = self._get_raw_value(cells_list, 'Time')
        units = self._get_raw_value(cells_list, 'Units')

        # skip the row entirely if no in-use column has a value in it
        any_value = _has_value(line_or_assay_name, mtype, val, time, units)
        if not any_value:
            return None

        # now that we've seen at least a single value in the row, do more rigorous parsing /
        # verification of the values
        line_or_assay_name = self._parse_and_verify_val(line_or_assay_name, row_index, 'Line Name')
        mtype = self._parse_and_verify_val(mtype, row_index, 'Measurement Type')
        val = self._parse_and_verify_val(val, row_index, 'Value')
        time = self._parse_and_verify_val(time, row_index, 'Time')
        units = self._parse_and_verify_val(units, row_index, 'Units')

        if isinstance(val, collections.Iterable):
            data = [time]
            for item in val:
                data.append(item)
        else:
            data = [time, val]

        m = MeasurementParseRecord(
            line_or_assay_name=line_or_assay_name,
            mtype_name=mtype,
            data=data,
            units_name=units,
            src_id=f'row {row_index+1}'
        )
        self._measurements.append(m)

        # track unique observed values
        if line_or_assay_name:
            self.unique_line_or_assay_names.add(line_or_assay_name)
        if val:
            self.unique_mtypes.add(mtype)
        if units:
            self.unique_units.add(units)

    @property
    def mtypes(self):
        return self.unique_mtypes

    @property
    def line_or_assay_names(self):
        return self.unique_line_or_assay_names

    @property
    def units(self):
        return self.unique_units

    @property
    def series_data(self):
        return self._measurements

    @property
    def measurement_count(self):
        return len(self._measurements)

    @property
    def has_all_times(self):
        """
        Tests whether the parsed file contained time values for all measurements.
        """
        return True  # overrides default False.  Time is a required column in this format.

    def has_all_units(self):
        """
        Tests whether the parsed file contained units for all measurements
        Partial units should be treated as a parse error.
        """
        return True  # overrides default False.  Units is a required column in this format
